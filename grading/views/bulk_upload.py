"""
Bulk Grade Upload Views

Handles uploading student grades from Excel files.
"""

from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from copy import copy

from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from rest_framework import status as http_status
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from ..access_policies import GradebookAccessPolicy

from academics.models import Section, AcademicYear, MarkingPeriod
from common.utils import get_object_by_uuid_or_fields
from students.models import Student
from grading.models import Assessment, Grade

class BulkGradeUploadView(APIView):
    """
    POST /sections/<section_id>/grades/upload/
    
    Upload student grades from Excel file.
    
    Query Parameters:
    - override_grades: Boolean (optional, default=false) - if true, updates grades regardless of status; 
                       if false, only updates grades with 'draft' or null status
    
    Excel Format (Metadata in first rows, then student data):
    
    Row 1: Grade Level: | <grade_level_name>
    Row 2: Section: | <section_name>
    Row 3: Subject Code: | <subject_code_or_name>
    Row 4: Academic Year: | <academic_year>
    Row 5: Marking Period Code: | <marking_period_code_or_name>
    Row 6: (blank)
    Row 7: Instructions: Do NOT modify...
    Row 8: (blank)
    Row 9: Student ID | Student Name | <Assessment1> | <Assessment2> | ...
    Row 10+: Student data rows
    
    Example:
    Grade Level:	Nursery 1
    Section:	General
    Subject Code:	ART
    Academic Year:	2025-2026
    Marking Period Code:	MP6
    
    Instructions: Do NOT modify the metadata above or student information columns below. Only enter scores in the assessment columns.
    
    Student ID	Student Name	Assignment	Attendance	Participation	Quiz	Test
    0121774	Michael Ashley Blair	5	5	5	20	40
    0121781	Brian Andrea Diaz	3	5	5	20	40
    """
    
    parser_classes = (MultiPartParser, FormParser)

    @staticmethod
    def _parse_bool(value, default=False):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    @staticmethod
    def _normalize_blank_condition_status(raw_value):
        """Normalize blank-cell upload behavior to supported condition statuses."""
        if raw_value in [None, ""]:
            return Grade.ConditionStatus.NO_GRADE

        normalized = str(raw_value).strip().lower().replace(" ", "_")

        if normalized in {"ng", "no-grade", "nograde"}:
            return Grade.ConditionStatus.NO_GRADE

        if normalized in {"missing", "excused", "absent", "not_submitted", "withdrawn"}:
            return Grade.ConditionStatus.NO_GRADE

        allowed = {
            Grade.ConditionStatus.NO_GRADE,
            Grade.ConditionStatus.INCOMPLETE,
            Grade.ConditionStatus.EXEMPT,
            Grade.ConditionStatus.PENDING,
        }
        return normalized if normalized in allowed else Grade.ConditionStatus.NO_GRADE
    
    @transaction.atomic
    def post(self, request, section_id):
        # Read override flag from query params first, then form body for robustness.
        override_raw = request.query_params.get('override_grades')
        if override_raw is None:
            override_raw = request.data.get('override_grades')
        override_grades = self._parse_bool(override_raw, default=False)

        # dry_run=true processes the file and computes all stats/errors but rolls
        # back the transaction so nothing is persisted.
        dry_run_raw = request.query_params.get('dry_run')
        if dry_run_raw is None:
            dry_run_raw = request.data.get('dry_run')
        dry_run = self._parse_bool(dry_run_raw, default=False)
        
        # Verify section exists
        section = get_object_or_404(Section, pk=section_id)
        
        # Get uploaded file
        if 'file' not in request.FILES:
            return Response({
                'detail': 'No file uploaded. Please provide an Excel file.'
            }, status=http_status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = request.FILES['file']
        
        # Validate file type
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'detail': 'Invalid file type. Please upload an Excel file (.xlsx or .xls).'
            }, status=http_status.HTTP_400_BAD_REQUEST)
        
        # Validate file size (max 10MB)
        max_file_size = 10 * 1024 * 1024  # 10MB in bytes
        if uploaded_file.size > max_file_size:
            return Response({
                'detail': f'File size exceeds maximum allowed size of 10MB. Your file is {uploaded_file.size / (1024 * 1024):.2f}MB.'
            }, status=http_status.HTTP_400_BAD_REQUEST)
        
        try:
            # Process the Excel file
            template_type = request.data.get('template_type', request.query_params.get('template_type', 'assessment_columns'))

            raw_blank_condition_status = request.data.get(
                'blank_condition_status',
                request.query_params.get('blank_condition_status')
            )
            blank_condition_status = self._normalize_blank_condition_status(raw_blank_condition_status)

            # grade_status lets the caller control what status newly created / updated grades receive.
            # Only a safe subset is accepted so callers cannot inject arbitrary values.
            _allowed_statuses = {
                Grade.Status.DRAFT,
                Grade.Status.PENDING,
                Grade.Status.SUBMITTED,
                Grade.Status.APPROVED,
            }
            raw_grade_status = str(request.data.get('grade_status', 'draft')).strip().lower()
            grade_status = raw_grade_status if raw_grade_status in _allowed_statuses else Grade.Status.DRAFT

            result = self._process_excel_file(
                uploaded_file,
                section,
                override_grades,
                request.user,
                template_type,
                grade_status,
                blank_condition_status,
            )

            if dry_run:
                # Roll back all DB writes; the caller just wants the preview stats.
                transaction.set_rollback(True)
                result['dry_run'] = True

            return Response(result, status=http_status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'detail': f'Error processing file: {str(e)}'
            }, status=http_status.HTTP_400_BAD_REQUEST)
    
    def _process_excel_file(
        self,
        file,
        section,
        override_grades,
        user,
        template_type='assessment_columns',
        grade_status='draft',
        blank_condition_status=Grade.ConditionStatus.NO_GRADE,
    ):
        """Process the uploaded Excel file and create/update grades"""
        import pandas as pd
        from academics.models import Subject, GradeLevel

        subject_has_code = any(field.name == 'code' for field in Subject._meta.fields)

        def _normalize_excel_text(value):
            """
            Normalize Excel cell values to trimmed strings without introducing
            numeric coercion artifacts (e.g. 20025.0).
            """
            if value is None:
                return ''

            if isinstance(value, str):
                return value.strip()

            try:
                if pd.isna(value):
                    return ''
            except Exception:
                pass

            if isinstance(value, float):
                # Excel often loads numeric IDs as float; strip only trailing .0.
                if value.is_integer():
                    return str(int(value))
                return format(value, 'f').rstrip('0').rstrip('.')

            if isinstance(value, int):
                return str(value)

            return str(value).strip()
        
        # ========================================================================
        # STEP 1: Read Excel file and extract metadata from first rows
        # ========================================================================
        try:
            # Read the entire Excel file without headers first
            df_raw = pd.read_excel(file, header=None)
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")
        
        if df_raw.empty:
            raise ValueError("Excel file is empty. Please provide data to upload.")
        
        # Extract metadata from first 5 rows
        metadata = {}
        try:
            # Parse metadata by labels so template layout can evolve safely.
            max_scan_rows = min(20, len(df_raw))
            label_to_row = {}
            for i in range(max_scan_rows):
                if len(df_raw.columns) == 0:
                    continue
                first_col = str(df_raw.iloc[i, 0]).strip().lower()
                if first_col.startswith('grade level'):
                    label_to_row['grade_level'] = i
                elif first_col.startswith('section'):
                    label_to_row['section'] = i
                elif first_col.startswith('subject code'):
                    label_to_row['subject'] = i
                elif first_col.startswith('subject name'):
                    label_to_row['subject_name'] = i
                elif first_col.startswith('academic year'):
                    label_to_row['academic_year'] = i
                elif first_col.startswith('marking period code'):
                    label_to_row['marking_period'] = i

            if 'grade_level' not in label_to_row:
                raise ValueError("Could not find 'Grade Level:' in the first 20 rows")

            def _value_at(row_idx):
                if row_idx is None or row_idx >= len(df_raw) or len(df_raw.columns) <= 1:
                    return None
                return str(df_raw.iloc[row_idx, 1]).strip()

            metadata['grade_level'] = _value_at(label_to_row.get('grade_level'))
            metadata['section'] = _value_at(label_to_row.get('section'))
            metadata['subject'] = _value_at(label_to_row.get('subject'))
            metadata['academic_year'] = _value_at(label_to_row.get('academic_year'))

            # Route based on the template_type parameter sent by the client.
            is_mp_columns = (template_type == 'marking_period_columns')

            if is_mp_columns:
                metadata['template_type'] = 'marking_period_columns'
                metadata['marking_period'] = None
            else:
                metadata['template_type'] = 'assessment_columns'
                metadata['marking_period'] = _value_at(label_to_row.get('marking_period'))

            # Detect header row by labels to support templates with optional Subject Name row.
            metadata['header_row'] = None
            header_scan_start = label_to_row.get('grade_level', 0)
            header_scan_end = min(header_scan_start + 25, len(df_raw))
            for i in range(header_scan_start, header_scan_end):
                row_values = [str(v).strip().lower() for v in df_raw.iloc[i].tolist()]
                if any(v in {'student id', 'id', 'id_number', 'student_id'} for v in row_values) and any(
                    v in {'student name', 'name', 'full_name', 'student_name'} for v in row_values
                ):
                    metadata['header_row'] = i
                    break

            if metadata['header_row'] is None:
                raise ValueError(
                    "Could not find a header row containing both Student ID and Student Name."
                )
            
        except Exception as e:
            raise ValueError(f"Failed to extract metadata from Excel file. Ensure metadata format is correct. Error: {str(e)}")
        
        # Validate metadata
        if not metadata.get('grade_level') or metadata['grade_level'] == 'nan':
            raise ValueError("Grade Level is missing in row 1. Format: 'Grade Level:' in column A, value in column B")
        
        if not metadata.get('section') or metadata['section'] == 'nan':
            raise ValueError("Section is missing in row 2. Format: 'Section:' in column A, value in column B")
        
        if not metadata.get('subject') or metadata['subject'] == 'nan':
            raise ValueError("Subject Code is missing in row 3. Format: 'Subject Code:' in column A, value in column B")
        
        if not metadata.get('academic_year') or metadata['academic_year'] == 'nan':
            raise ValueError("Academic Year is missing in row 4. Format: 'Academic Year:' in column A, value in column B")

        # Route to appropriate handler based on template type
        if metadata.get('template_type') == 'marking_period_columns':
            return self._process_marking_period_columns(
                file,
                df_raw,
                metadata,
                section,
                override_grades,
                user,
                grade_status,
                blank_condition_status,
            )

        # --- Original assessment_columns path ---
        if not metadata.get('marking_period') or metadata['marking_period'] == 'nan':
            raise ValueError("Marking Period Code is missing in row 5. Format: 'Marking Period Code:' in column A, value in column B")
        
        # ========================================================================
        # STEP 2: Validate metadata against database
        # ========================================================================
        
        # Validate Grade Level
        try:
            grade_level = GradeLevel.objects.get(name=metadata['grade_level'])
        except GradeLevel.DoesNotExist:
            raise ValueError(f"Grade Level '{metadata['grade_level']}' not found in the system")
        
        # Validate Section
        if section.name != metadata['section']:
            raise ValueError(f"Section mismatch. Expected '{section.name}' but file has '{metadata['section']}'")
        
        # Validate Academic Year
        try:
            academic_year = AcademicYear.objects.get(name=metadata['academic_year'])
        except AcademicYear.DoesNotExist:
            raise ValueError(f"Academic Year '{metadata['academic_year']}' not found in the system")
        
        # Validate Marking Period
        try:
            marking_period = MarkingPeriod.objects.get(
                Q(name__iexact=metadata['marking_period']) | Q(short_name__iexact=metadata['marking_period']),
                semester__academic_year=academic_year
            )
        except MarkingPeriod.DoesNotExist:
            raise ValueError(
                f"Marking Period '{metadata['marking_period']}' was not found by code or name for academic year '{metadata['academic_year']}'"
            )
        except MarkingPeriod.MultipleObjectsReturned:
            raise ValueError(
                f"Marking Period '{metadata['marking_period']}' is ambiguous for academic year '{metadata['academic_year']}'. Use a unique short name or name."
            )
        
        # ========================================================================
        # STEP 3: Read student data (headers row is dynamic based on metadata position)
        # ========================================================================
        try:
            # Use the header_row we calculated earlier
            header_row_index = metadata.get('header_row', 8)
            df = pd.read_excel(file, header=header_row_index, dtype=object)
            
            # Rename columns to standardize (handle possible column name variations)
            column_mapping = {}
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if col_lower in ['student id', 'id', 'id_number', 'student_id']:
                    column_mapping[col] = 'student_id'
                elif col_lower in ['student name', 'name', 'full_name', 'student_name']:
                    column_mapping[col] = 'student_name'
            
            df.rename(columns=column_mapping, inplace=True)

            # Treat IDs and names strictly as text to preserve leading zeros and
            # avoid float artifacts from Excel numeric cells.
            if 'student_id' in df.columns:
                df['student_id'] = df['student_id'].apply(_normalize_excel_text)
            if 'student_name' in df.columns:
                df['student_name'] = df['student_name'].apply(_normalize_excel_text)
            
        except Exception as e:
            raise ValueError(f"Failed to read student data from Excel file (header at row {header_row_index + 1}): {str(e)}")
        
        # Validate file is not empty
        if df.empty:
            raise ValueError(f"No student data found in Excel file. Student data should start from row {header_row_index + 2}.")
        
        # Validate required columns
        if 'student_id' not in df.columns:
            raise ValueError(f"'Student ID' or 'id_number' column not found in row {header_row_index + 1}. Ensure the header row has one of these columns.")

        # Get assessment columns (all columns except student_id and optional student_name)
        assessment_columns = [col for col in df.columns if col not in ['student_id', 'student_name']]
        
        if not assessment_columns:
            raise ValueError("No assessment columns found. Please add at least one assessment column after the student identifier columns.")

        # Validate Subject using the current upload context instead of a global name lookup.
        subject_identifier = metadata['subject']
        subject_lookup = Q(name=subject_identifier)
        if subject_has_code:
            subject_lookup |= Q(code=subject_identifier)

        subject_candidates = Subject.objects.filter(
            subject_lookup,
            gradebooks__section=section,
            gradebooks__academic_year=academic_year,
            gradebooks__assessments__marking_period=marking_period,
            gradebooks__assessments__name__in=assessment_columns,
            gradebooks__assessments__active=True,
        ).distinct()

        subject_count = subject_candidates.count()
        if subject_count == 0:
            raise ValueError(
                f"Subject '{metadata['subject']}' was not found by code or name for section '{section.name}' in academic year '{academic_year.name}'."
            )
        if subject_count > 1:
            raise ValueError(
                f"Subject '{metadata['subject']}' is ambiguous for this upload context. Please ensure the section, marking period, and assessment columns match a single gradebook."
            )

        subject = subject_candidates.first()
        
        # Statistics
        stats = {
            'total_rows': len(df),
            'students_processed': 0,
            'grades_created': 0,
            'grades_updated': 0,
            'grades_skipped': 0,
            'grades_locked': 0,
            'errors': [],
            'warnings': [],
            'metadata': metadata
        }
        
        # ========================================================================
        # STEP 4: Pre-fetch all needed data to reduce database queries
        # ========================================================================
        
        # Extract unique student IDs
        student_ids = df['student_id'].dropna().astype(str).str.strip().unique()
        
        # Pre-fetch students with the shared UUID-or-field resolver so id_number values
        # do not get sent through UUID filters and trigger validation errors.
        students_map = {}
        for student_lookup in student_ids:
            try:
                student = get_object_by_uuid_or_fields(
                    Student,
                    student_lookup,
                    fields=['id_number', 'prev_id_number'],
                )
            except Student.DoesNotExist:
                continue

            students_map[str(student_lookup)] = student
            students_map[str(student.id_number)] = student
            students_map[str(student.id)] = student

            prev_id_number = getattr(student, 'prev_id_number', None)
            if prev_id_number:
                students_map[str(prev_id_number)] = student
        
        # Pre-fetch assessments for this section, academic year, subject, and marking period
        assessments_cache = {}
        assessments_qs = Assessment.objects.filter(
            gradebook__section=section,
            gradebook__academic_year=academic_year,
            gradebook__section_subject__subject=subject,
            marking_period=marking_period,
            active=True
        ).select_related(
            'marking_period',
            'assessment_type',
            'gradebook',
            'gradebook__section_subject',
            'gradebook__section_subject__subject'
        )
        
        for assessment in assessments_qs:
            # Cache by assessment name for easy lookup
            assessments_cache[assessment.name] = assessment
        
        # Pre-fetch enrollments
        from students.models import Enrollment
        enrollments_cache = {}
        for enrollment in Enrollment.objects.filter(
            academic_year=academic_year,
            section=section,
            student__in=students_map.values()
        ).select_related('student'):
            enrollments_cache[enrollment.student.id] = enrollment.id
        
        # Pre-fetch existing grades
        existing_grades_cache = {}
        for grade in Grade.objects.filter(
            section=section,
            academic_year=academic_year,
            subject=subject,
            student__in=students_map.values(),
            assessment__marking_period=marking_period
        ).select_related('assessment', 'student'):
            key = (grade.assessment.id, grade.student.id)
            existing_grades_cache[key] = grade
        
        # ========================================================================
        # STEP 5: Process student rows
        # ========================================================================
        
        grades_to_create = []
        grades_to_update = []
        processed_students = set()
        
        # Get header row index for accurate row number reporting
        header_row_index = metadata.get('header_row', 8)
        
        for index, row in df.iterrows():
            # Excel row number (accounting for header row position + 1 for data rows)
            # Add 2: +1 for 0-indexed to 1-indexed, +1 because data starts after header
            row_number = header_row_index + index + 2
            
            try:
                # Get student ID
                student_id = str(row['student_id']).strip()
                if pd.isna(student_id) or not student_id:
                    stats['errors'].append({
                        'row': row_number,
                        'error': 'Student ID is missing'
                    })
                    continue
                
                # Lookup student
                student = students_map.get(student_id)
                if not student:
                    stats['errors'].append({
                        'row': row_number,
                        'student_id': student_id,
                        'error': f'Student not found with ID: {student_id}'
                    })
                    continue
                
                # Get enrollment
                enrollment_id = enrollments_cache.get(student.id)
                if not enrollment_id:
                    stats['errors'].append({
                        'row': row_number,
                        'student_id': student_id,
                        'error': f'Student is not enrolled in {section.name} for {academic_year.name}'
                    })
                    continue
                
                # Process each assessment column
                student_processed = False
                
                for assessment_col in assessment_columns:
                    score_value = row[assessment_col]

                    # Blank/null cells are treated as a configurable condition status.
                    is_blank_score = pd.isna(score_value) or str(score_value).strip() == ''
                    
                    # Lookup assessment
                    assessment = assessments_cache.get(assessment_col)
                    
                    if not assessment:
                        stats['errors'].append({
                            'row': row_number,
                            'student_id': student_id,
                            'assessment': assessment_col,
                            'error': f"Assessment '{assessment_col}' not found for {metadata['marking_period']}"
                        })
                        continue
                    
                    # Validate and convert score/condition payload.
                    if is_blank_score:
                        score = None
                        condition_status = blank_condition_status
                        condition_reason = None
                    else:
                        try:
                            score = Decimal(str(score_value).strip())

                            if score < 0:
                                stats['errors'].append({
                                    'row': row_number,
                                    'student_id': student_id,
                                    'assessment': assessment_col,
                                    'error': f'Score cannot be negative: {score}'
                                })
                                continue

                            if assessment.max_score and score > assessment.max_score:
                                stats['errors'].append({
                                    'row': row_number,
                                    'student_id': student_id,
                                    'assessment': assessment_col,
                                    'error': f'Score {score} exceeds maximum score of {assessment.max_score}'
                                })
                                continue

                            if score.as_tuple().exponent < -2:
                                stats['errors'].append({
                                    'row': row_number,
                                    'student_id': student_id,
                                    'assessment': assessment_col,
                                    'error': f'Score {score} has too many decimal places. Maximum 2 allowed.'
                                })
                                continue
                        except (InvalidOperation, ValueError, AttributeError):
                            stats['errors'].append({
                                'row': row_number,
                                'student_id': student_id,
                                'assessment': assessment_col,
                                'error': f'Invalid score value: {score_value}. Must be a number.'
                            })
                            continue

                        condition_status = Grade.ConditionStatus.GRADED
                        condition_reason = None
                    
                    # Check if grade exists
                    grade_key = (assessment.id, student.id)
                    existing_grade = existing_grades_cache.get(grade_key)
                    
                    if existing_grade:
                        same_score = existing_grade.score == score
                        same_condition = existing_grade.condition_status == condition_status

                        if same_score and same_condition:
                            stats['grades_skipped'] += 1
                            student_processed = True
                            continue

                        # Check if we can update
                        can_update = override_grades or existing_grade.status in [Grade.Status.DRAFT, None]
                        
                        if can_update:
                            existing_grade.score = score
                            existing_grade.condition_status = condition_status
                            existing_grade.condition_reason = condition_reason
                            existing_grade.status = grade_status
                            existing_grade.updated_by = user
                            grades_to_update.append(existing_grade)
                            stats['grades_updated'] += 1
                        else:
                            stats['grades_locked'] += 1
                            stats['warnings'].append({
                                'row': row_number,
                                'student_id': student_id,
                                'assessment': assessment_col,
                                'warning': f'Grade is {existing_grade.get_status_display()}. Use override_grades=true to update.'
                            })
                    else:
                        # Create new grade
                        new_grade = Grade(
                            assessment=assessment,
                            student=student,
                            score=score,
                            condition_status=condition_status,
                            condition_reason=condition_reason,
                            status=grade_status,
                            enrollment_id=enrollment_id,
                            academic_year=academic_year,
                            section=section,
                            subject=subject,
                            created_by=user,
                            updated_by=user
                        )
                        grades_to_create.append(new_grade)
                        # Update cache to prevent duplicates
                        existing_grades_cache[grade_key] = new_grade
                        stats['grades_created'] += 1
                    
                    student_processed = True
                
                if student_processed:
                    processed_students.add(student.id)
                    
            except Exception as e:
                stats['errors'].append({
                    'row': row_number,
                    'error': f'Unexpected error: {str(e)}'
                })
                continue
        
        # ========================================================================
        # STEP 6: Bulk create/update grades
        # ========================================================================
        
        if grades_to_create:
            Grade.objects.bulk_create(grades_to_create, batch_size=500)
        
        if grades_to_update:
            Grade.objects.bulk_update(
                grades_to_update, 
                ['score', 'condition_status', 'condition_reason', 'status', 'updated_by', 'updated_at'],
                batch_size=500
            )
        
        stats['students_processed'] = len(processed_students)
        
        # Build response message
        message = f"Successfully processed {stats['students_processed']} students for {metadata['subject']} - {metadata['marking_period']}. "
        message += f"Created {stats['grades_created']} new grades, updated {stats['grades_updated']} existing grades."
        
        if stats['grades_locked'] > 0:
            message += f" {stats['grades_locked']} grades were locked (use override_grades=true to update)."
        
        if stats['warnings']:
            message += f" {len(stats['warnings'])} warnings occurred."
        
        if stats['errors']:
            message += f" {len(stats['errors'])} errors occurred."
        
        return {
            'detail': message,
            'metadata': {
                **metadata,
                'blank_condition_status': blank_condition_status,
            },
            'statistics': {
                'total_rows': stats['total_rows'],
                'students_processed': stats['students_processed'],
                'grades_created': stats['grades_created'],
                'grades_updated': stats['grades_updated'],
                'grades_locked': stats['grades_locked'],
                'warning_count': len(stats['warnings']),
                'error_count': len(stats['errors'])
            },
            'warnings': stats['warnings'][:50],
            'errors': stats['errors'][:50]
        }

    @transaction.atomic
    def _process_marking_period_columns(
        self,
        file,
        df_raw,
        metadata,
        section,
        override_grades,
        user,
        grade_status='draft',
        blank_condition_status=Grade.ConditionStatus.NO_GRADE,
    ):
        """
        Process Template 2: Marking Period Columns format.

        Layout:
          Row 1: Grade Level:  | <value>
          Row 2: Section:      | <value>
          Row 3: Subject Code: | <value>
          Row 4: Academic Year:| <value>
          Row 5: Template Type:| marking_period_columns
          Row 6: (blank)
          Row 7: Instructions: | ...
          Row 8: (blank)
          Row 9: Student ID | Student Name | <MP short_name1> | <MP short_name2> | ...
          Row 10+: data rows

        Each marking period column holds the single final grade for that student/period.
        Grades are SKIPPED if the incoming value equals the existing score.
        Grades are UPDATED if the incoming value differs (respecting override_grades for locked grades).
        Grades are CREATED if none exist yet.
        """
        import pandas as pd
        from academics.models import Subject, GradeLevel

        subject_has_code = any(field.name == 'code' for field in Subject._meta.fields)

        def _normalize_excel_text(value):
            if value is None:
                return ''
            if isinstance(value, str):
                return value.strip()
            try:
                if pd.isna(value):
                    return ''
            except Exception:
                pass
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                return format(value, 'f').rstrip('0').rstrip('.')
            if isinstance(value, int):
                return str(value)
            return str(value).strip()

        # =====================================================================
        # STEP 1: Validate common metadata
        # =====================================================================

        try:
            GradeLevel.objects.get(name=metadata['grade_level'])
        except GradeLevel.DoesNotExist:
            raise ValueError(f"Grade Level '{metadata['grade_level']}' not found in the system")

        if section.name != metadata['section']:
            raise ValueError(f"Section mismatch. Expected '{section.name}' but file has '{metadata['section']}'")

        try:
            academic_year = AcademicYear.objects.get(name=metadata['academic_year'])
        except AcademicYear.DoesNotExist:
            raise ValueError(f"Academic Year '{metadata['academic_year']}' not found in the system")

        # =====================================================================
        # STEP 2: Read student data rows
        # =====================================================================
        header_row_index = metadata.get('header_row', 7)
        try:
            df = pd.read_excel(file, header=header_row_index, dtype=object)
        except Exception as e:
            raise ValueError(f"Failed to read student data: {str(e)}")

        if df.empty:
            raise ValueError(f"No student data found. Student data should start from row {header_row_index + 2}.")

        # Normalize student id/name columns
        column_mapping = {}
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if col_lower in ['student id', 'id', 'id_number', 'student_id']:
                column_mapping[col] = 'student_id'
            elif col_lower in ['student name', 'name', 'full_name', 'student_name']:
                column_mapping[col] = 'student_name'
        df.rename(columns=column_mapping, inplace=True)

        if 'student_id' in df.columns:
            df['student_id'] = df['student_id'].apply(_normalize_excel_text)
        if 'student_name' in df.columns:
            df['student_name'] = df['student_name'].apply(_normalize_excel_text)

        if 'student_id' not in df.columns:
            raise ValueError("'Student ID' column not found in header row.")

        # Columns after student_id and student_name are marking period labels.
        # Skip average/summary columns (sem avg, yearly avg) — they are read-only in the template.
        mp_columns = [
            col for col in df.columns
            if col not in ['student_id', 'student_name']
            and 'avg' not in str(col).strip().lower()
            and 'average' not in str(col).strip().lower()
        ]
        if not mp_columns:
            raise ValueError("No marking period columns found after the student identifier columns.")

        # =====================================================================
        # STEP 3: Resolve subject
        # =====================================================================
        subject_identifier = metadata['subject']
        subject_lookup = Q(name=subject_identifier)
        if subject_has_code:
            subject_lookup |= Q(code=subject_identifier)

        subject_candidates = Subject.objects.filter(
            subject_lookup,
            gradebooks__section=section,
            gradebooks__academic_year=academic_year,
        ).distinct()

        subject_count = subject_candidates.count()
        if subject_count == 0:
            raise ValueError(
                f"Subject '{subject_identifier}' was not found for section '{section.name}' "
                f"in academic year '{academic_year.name}'."
            )
        if subject_count > 1:
            raise ValueError(
                f"Subject '{subject_identifier}' is ambiguous. Multiple gradebooks matched."
            )
        subject = subject_candidates.first()

        # =====================================================================
        # STEP 4: Resolve marking periods and their single-entry assessments
        # =====================================================================
        # Build map: mp_label (short_name or name) -> MarkingPeriod  (case-insensitive keys)
        all_mps = MarkingPeriod.objects.filter(
            semester__academic_year=academic_year
        ).select_related('semester')
        mp_by_label = {}
        for mp in all_mps:
            mp_by_label[mp.short_name.strip().lower()] = mp
            mp_by_label[mp.name.strip().lower()] = mp

        # For each MP column header, resolve the MarkingPeriod and its single-entry assessment
        mp_assessment_map = {}  # label -> (MarkingPeriod, Assessment | None)
        unresolved_mp_columns = []
        for col_label in mp_columns:
            col_str = str(col_label).strip()
            mp = mp_by_label.get(col_str.lower())
            if not mp:
                unresolved_mp_columns.append(col_str)
                continue
            # Find the single-entry assessment for this marking period in the gradebook
            assessment_qs = Assessment.objects.filter(
                gradebook__section=section,
                gradebook__academic_year=academic_year,
                gradebook__section_subject__subject=subject,
                marking_period=mp,
                assessment_type__is_single_entry=True,
                active=True,
            ).select_related('assessment_type', 'marking_period', 'gradebook')
            assessment = assessment_qs.first()
            mp_assessment_map[col_label] = (mp, assessment)

        if unresolved_mp_columns:
            raise ValueError(
                f"These marking period columns could not be matched by short name or name: "
                f"{', '.join(unresolved_mp_columns)}. "
                f"Ensure column headers exactly match marking period short names or names."
            )

        # =====================================================================
        # STEP 5: Pre-fetch students, enrollments, existing grades
        # =====================================================================
        from students.models import Enrollment

        student_ids_raw = df['student_id'].dropna().astype(str).str.strip().unique()
        students_map = {}
        for sid in student_ids_raw:
            try:
                student = get_object_by_uuid_or_fields(
                    Student, sid, fields=['id_number', 'prev_id_number'],
                )
            except Student.DoesNotExist:
                continue
            for key in [str(sid), str(student.id_number), str(student.id)]:
                students_map[key] = student
            if getattr(student, 'prev_id_number', None):
                students_map[str(student.prev_id_number)] = student

        # All assessments for this subject/section/year across all marking periods
        all_assessments = list(
            Assessment.objects.filter(
                gradebook__section=section,
                gradebook__academic_year=academic_year,
                gradebook__section_subject__subject=subject,
                assessment_type__is_single_entry=True,
                active=True,
            ).select_related('marking_period')
        )
        assessment_ids = [a.id for a in all_assessments]

        enrollments_cache = {}
        for enr in Enrollment.objects.filter(
            academic_year=academic_year,
            section=section,
            student__in=students_map.values(),
        ).select_related('student'):
            enrollments_cache[enr.student.id] = enr.id

        existing_grades_cache = {}
        for grade in Grade.objects.filter(
            section=section,
            academic_year=academic_year,
            subject=subject,
            student__in=students_map.values(),
            assessment__id__in=assessment_ids,
        ).select_related('assessment', 'student'):
            existing_grades_cache[(grade.assessment.id, grade.student.id)] = grade

        # =====================================================================
        # STEP 6: Process rows
        # =====================================================================
        stats = {
            'total_rows': len(df),
            'students_processed': 0,
            'grades_created': 0,
            'grades_updated': 0,
            'grades_skipped': 0,
            'grades_locked': 0,
            'errors': [],
            'warnings': [],
        }

        grades_to_create = []
        grades_to_update = []
        processed_students = set()

        for index, row in df.iterrows():
            row_number = header_row_index + index + 2
            try:
                student_id = str(row['student_id']).strip()
                if pd.isna(student_id) or not student_id:
                    stats['errors'].append({'row': row_number, 'error': 'Student ID is missing'})
                    continue

                student = students_map.get(student_id)
                if not student:
                    stats['errors'].append({
                        'row': row_number, 'student_id': student_id,
                        'error': f'Student not found with ID: {student_id}',
                    })
                    continue

                enrollment_id = enrollments_cache.get(student.id)
                if not enrollment_id:
                    stats['errors'].append({
                        'row': row_number, 'student_id': student_id,
                        'error': f'Student is not enrolled in {section.name} for {academic_year.name}',
                    })
                    continue

                student_processed = False

                for col_label, (mp, assessment) in mp_assessment_map.items():
                    score_value = row[col_label]

                    # Blank/null cells are treated as a configurable condition status.
                    try:
                        is_blank = pd.isna(score_value) or str(score_value).strip() == ''
                    except Exception:
                        is_blank = not str(score_value).strip()

                    if assessment is None:
                        stats['errors'].append({
                            'row': row_number, 'student_id': student_id,
                            'error': f"No single-entry assessment found for marking period '{col_label}'. "
                                     f"Ensure a single-entry assessment exists for this subject and period.",
                        })
                        continue

                    if is_blank:
                        score = None
                        condition_status = blank_condition_status
                        condition_reason = None
                    else:
                        # Parse score
                        try:
                            score = Decimal(str(score_value).strip())
                            if score < 0:
                                stats['errors'].append({
                                    'row': row_number, 'student_id': student_id,
                                    'error': f'[{col_label}] Score cannot be negative: {score}',
                                })
                                continue
                            if assessment.max_score and score > assessment.max_score:
                                stats['errors'].append({
                                    'row': row_number, 'student_id': student_id,
                                    'error': f'[{col_label}] Score {score} exceeds maximum of {assessment.max_score}',
                                })
                                continue
                            if score.as_tuple().exponent < -2:
                                stats['errors'].append({
                                    'row': row_number, 'student_id': student_id,
                                    'error': f'[{col_label}] Score {score} has too many decimal places (max 2).',
                                })
                                continue
                        except (InvalidOperation, ValueError):
                            stats['errors'].append({
                                'row': row_number, 'student_id': student_id,
                                'error': f'[{col_label}] Invalid score value: {score_value}',
                            })
                            continue

                        condition_status = Grade.ConditionStatus.GRADED
                        condition_reason = None

                    grade_key = (assessment.id, student.id)
                    existing_grade = existing_grades_cache.get(grade_key)

                    if existing_grade:
                        # Preserve no-op optimization: skip if incoming score
                        # and condition status match existing grade.
                        if (
                            existing_grade.score == score
                            and existing_grade.condition_status == condition_status
                        ):
                            stats['grades_skipped'] += 1
                            student_processed = True
                            continue

                        can_update = override_grades or existing_grade.status in [Grade.Status.DRAFT, None]
                        if can_update:
                            existing_grade.score = score
                            existing_grade.condition_status = condition_status
                            existing_grade.condition_reason = condition_reason
                            existing_grade.status = grade_status
                            existing_grade.updated_by = user
                            grades_to_update.append(existing_grade)
                            stats['grades_updated'] += 1
                        else:
                            stats['grades_locked'] += 1
                            stats['warnings'].append({
                                'row': row_number, 'student_id': student_id,
                                'warning': f'[{col_label}] Grade is {existing_grade.get_status_display()}. Use override_grades=true to update.',
                            })
                    else:
                        new_grade = Grade(
                            assessment=assessment,
                            student=student,
                            score=score,
                            condition_status=condition_status,
                            condition_reason=condition_reason,
                            status=grade_status,
                            enrollment_id=enrollment_id,
                            academic_year=academic_year,
                            section=section,
                            subject=subject,
                            created_by=user,
                            updated_by=user,
                        )
                        grades_to_create.append(new_grade)
                        existing_grades_cache[grade_key] = new_grade
                        stats['grades_created'] += 1

                    student_processed = True

                if student_processed:
                    processed_students.add(student.id)

            except Exception as e:
                stats['errors'].append({'row': row_number, 'error': f'Unexpected error: {str(e)}'})
                continue

        # =====================================================================
        # STEP 7: Bulk write
        # =====================================================================
        if grades_to_create:
            Grade.objects.bulk_create(grades_to_create, batch_size=500)
        if grades_to_update:
            Grade.objects.bulk_update(
                grades_to_update,
                ['score', 'condition_status', 'condition_reason', 'status', 'updated_by', 'updated_at'],
                batch_size=500,
            )

        stats['students_processed'] = len(processed_students)

        message = (
            f"Marking period upload processed {stats['students_processed']} students for "
            f"{metadata['subject']} across {len(mp_columns)} marking period(s). "
            f"Created {stats['grades_created']}, updated {stats['grades_updated']}, "
            f"skipped {stats['grades_skipped']} unchanged."
        )
        if stats['grades_locked'] > 0:
            message += f" {stats['grades_locked']} locked (use override_grades=true to update)."
        if stats['warnings']:
            message += f" {len(stats['warnings'])} warnings."
        if stats['errors']:
            message += f" {len(stats['errors'])} errors."

        return {
            'detail': message,
            'metadata': {
                **metadata,
                'template_type': 'marking_period_columns',
                'blank_condition_status': blank_condition_status,
            },
            'statistics': {
                'total_rows': stats['total_rows'],
                'students_processed': stats['students_processed'],
                'grades_created': stats['grades_created'],
                'grades_updated': stats['grades_updated'],
                'grades_skipped': stats['grades_skipped'],
                'grades_locked': stats['grades_locked'],
                'warning_count': len(stats['warnings']),
                'error_count': len(stats['errors']),
            },
            'warnings': stats['warnings'][:50],
            'errors': stats['errors'][:50],
        }


# ---------------------------------------------------------------------------
# Shared helpers for marking-period template workbooks (single- and multi-sheet)
# ---------------------------------------------------------------------------

_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / 'templates'
    / 'grades_upload_marking_periods.xlsx'
)


def _normalize_cell_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        trimmed = value.strip()
        return trimmed if trimmed else None
    return value


def _copy_row_style(worksheet, src_row: int, dest_row: int, col_count: int):
    for col_idx in range(1, col_count + 1):
        src_cell = worksheet.cell(src_row, col_idx)
        dest_cell = worksheet.cell(dest_row, col_idx)
        if src_cell.has_style:
            dest_cell._style = copy(src_cell._style)
        if src_cell.number_format:
            dest_cell.number_format = src_cell.number_format
        if src_cell.protection:
            dest_cell.protection = copy(src_cell.protection)
        if src_cell.alignment:
            dest_cell.alignment = copy(src_cell.alignment)

    src_dim = worksheet.row_dimensions.get(src_row)
    if src_dim and src_dim.height is not None:
        worksheet.row_dimensions[dest_row].height = src_dim.height


def _populate_marking_period_sheet(
    worksheet,
    *,
    grade_level: str,
    section_name: str,
    subject_code: str,
    subject_name: str,
    academic_year: str,
    headers: list,
    rows: list,
):
    """Populate a worksheet (pre-styled from the bundled template) with a
    marking-period grade template payload.

    Layout (aligned with the bundled template file):
      rows 1-5 : metadata label/value pairs
      row 7    : column headers
      row 8+   : student data rows
    """
    metadata_values = [
        ('Grade Level:', grade_level),
        ('Section:', section_name),
        ('Subject Code:', subject_code),
        ('Subject Name:', subject_name),
        ('Academic Year:', academic_year),
    ]

    for row_index, (label, value) in enumerate(metadata_values, start=1):
        worksheet.cell(row_index, 1).value = label
        worksheet.cell(row_index, 2).value = value

    header_row_index = 7
    data_start_row_index = header_row_index + 1

    template_data_style_row = worksheet.max_row
    if template_data_style_row < data_start_row_index:
        template_data_style_row = data_start_row_index

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(header_row_index, column_index).value = header

    max_col = max(len(headers), max((len(r) for r in rows), default=0))

    for row_offset, row_values in enumerate(rows):
        row_index = data_start_row_index + row_offset

        if row_index > template_data_style_row and max_col > 0:
            _copy_row_style(worksheet, template_data_style_row, row_index, max_col)

        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row_index, column_index).value = _normalize_cell_value(value)

    # Color-code grade values in all numeric grade columns:
    # <70 => red, >=70 => blue.
    if max_col >= 3 and rows:
        last_data_row = data_start_row_index + len(rows) - 1
        start_col_letter = worksheet.cell(data_start_row_index, 3).column_letter
        end_col_letter = worksheet.cell(data_start_row_index, max_col).column_letter
        grade_range = f"{start_col_letter}{data_start_row_index}:{end_col_letter}{last_data_row}"

        worksheet.conditional_formatting.add(
            grade_range,
            CellIsRule(operator='lessThan', formula=['70'], font=Font(color='FFDC2626')),
        )
        worksheet.conditional_formatting.add(
            grade_range,
            CellIsRule(operator='greaterThanOrEqual', formula=['70'], font=Font(color='FF2563EB')),
        )


_INVALID_SHEET_CHARS = str.maketrans({c: '_' for c in r'[]:*?/\\'})


def _sanitize_sheet_title(raw: str, used: set, fallback: str) -> str:
    """Sanitize an Excel worksheet title (<=31 chars, no forbidden chars,
    unique within *used*). *used* is updated in place with the returned title.
    """
    base = (raw or '').strip().translate(_INVALID_SHEET_CHARS)
    if not base:
        base = (fallback or 'Sheet').strip() or 'Sheet'
    base = base[:31]

    candidate = base
    counter = 2
    while candidate in used or not candidate:
        suffix = f' ({counter})'
        candidate = f'{base[: 31 - len(suffix)]}{suffix}'
        counter += 1
    used.add(candidate)
    return candidate


class BulkGradeTemplateDownloadView(APIView):
    """Download a styled bulk grade template using the bundled workbook file."""

    parser_classes = (JSONParser,)

    def post(self, request, section_id):
        section = get_object_or_404(Section, pk=section_id)

        template_type = str(request.data.get('template_type', 'marking_period_columns')).strip()
        if template_type != 'marking_period_columns':
            return Response(
                {'detail': 'Only marking_period_columns templates are supported by this endpoint.'},
                status=http_status.HTTP_400_BAD_REQUEST,
            )

        headers = request.data.get('headers') or []
        rows = request.data.get('rows') or []
        grade_level = str(request.data.get('grade_level', '')).strip()
        section_name = str(request.data.get('section') or section.name).strip()
        subject = str(request.data.get('subject', '')).strip()
        subject_name = str(request.data.get('subject_name', '')).strip()
        academic_year = str(request.data.get('academic_year', '')).strip()

        if not isinstance(headers, list) or not all(isinstance(header, str) for header in headers):
            return Response({'detail': 'headers must be an array of strings.'}, status=http_status.HTTP_400_BAD_REQUEST)
        if not isinstance(rows, list):
            return Response({'detail': 'rows must be an array.'}, status=http_status.HTTP_400_BAD_REQUEST)

        if not _TEMPLATE_PATH.exists():
            return Response(
                {'detail': 'Workbook template file was not found on the server.'},
                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        workbook = load_workbook(_TEMPLATE_PATH)
        worksheet = workbook[workbook.sheetnames[0]]

        _populate_marking_period_sheet(
            worksheet,
            grade_level=grade_level,
            section_name=section_name,
            subject_code=subject,
            subject_name=subject_name,
            academic_year=academic_year,
            headers=headers,
            rows=rows,
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="grades_upload_marking_periods.xlsx"'
        response['Cache-Control'] = 'no-store'
        return response


class BulkGradeAllSubjectsTemplateDownloadView(APIView):
    """
    POST /sections/<section_id>/grades-upload/template/all/

    Download a single Excel workbook that contains one pre-populated
    marking-period template sheet **per subject** for the given section.

    Body (all optional):
      - academic_year_id      : UUID of the academic year (default: current year)
      - include_grades        : bool. When true, populate each sheet with the
                                approved / filtered marking-period grades.
      - grade_status_filter   : Grade.status to include when include_grades is
                                true. Use 'all' (or omit) to include every
                                status.

    Each sheet layout:
      - metadata rows (grade level, section, subject code/name, academic year)
      - header row: Student ID | Student Name | <MP short names...> | Sem_N Avg
      - one row per enrolled student (id_number + "Last, First Middle" + grades)
    """

    parser_classes = (JSONParser,)

    def post(self, request, section_id):
        from students.models import Enrollment
        from grading.models import GradeBook
        from grading.utils import calculate_marking_period_percentage

        section = get_object_or_404(
            Section.objects.select_related('grade_level'),
            pk=section_id,
        )

        # Resolve academic year (defaults to current).
        academic_year_id = str(request.data.get('academic_year_id', '')).strip() or None
        if academic_year_id:
            try:
                academic_year = AcademicYear.objects.get(pk=academic_year_id)
            except AcademicYear.DoesNotExist:
                return Response(
                    {'detail': 'Academic year not found.'},
                    status=http_status.HTTP_404_NOT_FOUND,
                )
        else:
            academic_year = AcademicYear.get_current_academic_year()
            if academic_year is None:
                return Response(
                    {'detail': 'No current academic year is configured.'},
                    status=http_status.HTTP_400_BAD_REQUEST,
                )

        include_grades = bool(request.data.get('include_grades', False))
        raw_status_filter = str(request.data.get('grade_status_filter', 'all')).strip().lower()
        status_filter = 'any' if raw_status_filter in ('', 'all') else raw_status_filter

        # All gradebooks for this section+year, ordered by subject.
        gradebooks = list(
            GradeBook.objects.filter(section=section, academic_year=academic_year)
            .select_related('subject')
            .order_by('subject__code', 'subject__name')
        )

        if not gradebooks:
            return Response(
                {'detail': 'No gradebooks found for this section and academic year.'},
                status=http_status.HTTP_404_NOT_FOUND,
            )

        # Enrolled students in this section for the given year.
        enrollments = (
            Enrollment.objects.filter(section=section, academic_year=academic_year)
            .select_related('student')
            .order_by('student__last_name', 'student__first_name', 'student__middle_name')
        )
        students = [e.student for e in enrollments]

        # Marking periods for the year, grouped by semester (start_date order).
        marking_periods = list(
            MarkingPeriod.objects
            .filter(semester__academic_year=academic_year)
            .select_related('semester')
            .order_by('semester__start_date', 'start_date')
        )

        # Group marking periods by semester while preserving encounter order.
        semester_groups = []  # list of (semester_id_or_None, semester_name, [periods])
        seen = {}
        for mp in marking_periods:
            sem_id = mp.semester_id if mp.semester_id else None
            key = sem_id if sem_id is not None else '__no_semester__'
            if key not in seen:
                seen[key] = len(semester_groups)
                semester_groups.append((
                    key,
                    mp.semester.name if mp.semester else '',
                    [],
                ))
            semester_groups[seen[key]][2].append(mp)

        # Build the ordered column-header list (matches the frontend layout).
        headers = ['Student ID', 'Student Name']
        for idx, (_key, _name, periods) in enumerate(semester_groups, start=1):
            for mp in periods:
                headers.append(mp.short_name or mp.name)
            headers.append(f'Sem{idx} Avg')
        if len(semester_groups) > 1:
            headers.append('Yrly Avg')

        # Number of blank grade columns per student row.
        blank_grade_cell_count = len(headers) - 2  # minus Student ID + Name

        def _avg_if_complete(scores, expected_count):
            if expected_count == 0:
                return ''
            numeric = [s for s in scores if s is not None]
            if len(numeric) < expected_count:
                return ''
            return round(sum(numeric) / len(numeric))

        def _build_rows_for_gradebook(gradebook):
            rows = []
            for student in students:
                id_number = (student.id_number or '').strip()
                first = (student.first_name or '').strip()
                middle = (student.middle_name or '').strip()
                last = (student.last_name or '').strip()
                display_name = f"{last}, {first} {middle}".strip()
                while '  ' in display_name:
                    display_name = display_name.replace('  ', ' ')
                display_name = display_name.rstrip(',').strip()

                if not include_grades:
                    rows.append(
                        [id_number, display_name] + [''] * blank_grade_cell_count
                    )
                    continue

                sem_scores = {key: [] for key, _n, _p in semester_groups}
                sem_expected = {
                    key: len(periods) for key, _n, periods in semester_groups
                }
                cells = [id_number, display_name]

                for sem_idx, (sem_key, _sem_name, periods) in enumerate(
                    semester_groups, start=1
                ):
                    for mp in periods:
                        pct = calculate_marking_period_percentage(
                            gradebook, student, mp, status=status_filter,
                        )
                        score = int(round(float(pct))) if pct is not None else None
                        cells.append(score if score is not None else '')
                        sem_scores[sem_key].append(score)
                    cells.append(
                        _avg_if_complete(sem_scores[sem_key], sem_expected[sem_key])
                    )

                if len(semester_groups) > 1:
                    all_scores = [s for lst in sem_scores.values() for s in lst]
                    total_expected = sum(sem_expected.values())
                    cells.append(_avg_if_complete(all_scores, total_expected))

                rows.append(cells)
            return rows

        # If there are no enrolled students, emit an empty template row per
        # sheet so the workbook remains usable.
        empty_row = ['', ''] + [''] * blank_grade_cell_count

        if not _TEMPLATE_PATH.exists():
            return Response(
                {'detail': 'Workbook template file was not found on the server.'},
                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        workbook = load_workbook(_TEMPLATE_PATH)
        base_sheet = workbook[workbook.sheetnames[0]]

        grade_level_name = section.grade_level.name if section.grade_level_id else ''
        used_titles = set()

        for gradebook in gradebooks:
            subject = gradebook.subject
            subject_code = (subject.code or '').strip() if subject else ''
            subject_name = (subject.name or '').strip() if subject else ''

            worksheet = workbook.copy_worksheet(base_sheet)
            worksheet.title = _sanitize_sheet_title(
                subject_code or subject_name,
                used_titles,
                fallback='Subject',
            )

            rows = _build_rows_for_gradebook(gradebook) if students else [empty_row]

            _populate_marking_period_sheet(
                worksheet,
                grade_level=grade_level_name,
                section_name=section.name,
                subject_code=subject_code,
                subject_name=subject_name,
                academic_year=str(academic_year),
                headers=headers,
                rows=rows,
            )

        # Remove the untouched base sheet so only per-subject sheets remain.
        workbook.remove(base_sheet)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        section_part = ''.join(
            ch if ch.isalnum() else '_' for ch in (section.name or '')
        ).strip('_') or 'section'
        filename = f'grades_upload_{section_part}_all_subjects.xlsx'

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Cache-Control'] = 'no-store'
        return response
