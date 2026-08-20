from __future__ import annotations

import html
import io
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common.services.pdf_components import (
    append_pdf_document_header,
    append_pdf_subtitle,
    get_pdf_header_styles,
    resolve_tenant_school,
)
from reports.utils.export_helpers import (
    apply_xlsx_cell_style,
    format_amount_display,
    format_percentage_display,
    get_export_format,
    prepare_xlsx_numeric_value,
)


BLUE = "1F4E79"
LIGHT_BLUE = "D9EAF7"


def _display_date(value):
    if not value:
        return ""
    try:
        return datetime.fromisoformat(str(value)).strftime("%b %d, %Y")
    except ValueError:
        return str(value)


def _amount(value):
    if value is None:
        return "-"
    return format_amount_display(value)


def _paragraph(value, style):
    from reportlab.platypus import Paragraph

    return Paragraph(html.escape(str(value or "")), style)


def _pdf_table(headers, rows, body_style, *, widths=None, font_size=7):
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph, Table, TableStyle

    data = [[Paragraph(f"<b>{html.escape(str(header))}</b>", body_style) for header in headers]]
    data.extend([[_paragraph(cell, body_style) for cell in row] for row in rows])
    table = Table(data, repeatRows=1, colWidths=widths, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{LIGHT_BLUE}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#243B53")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("LEADING", (0, 0), (-1, -1), font_size + 2),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C2CC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
    ]))
    return table


def build_comprehensive_budget_pdf(request, payload, filename):  # noqa: C901
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    details = payload["details"]
    budget = details["budget"]
    currency = budget["currency"]
    buf = io.BytesIO()
    page_size = landscape(letter)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.3 * inch,
        bottomMargin=0.3 * inch,
    )
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle(
        "BudgetBody",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
    )
    section_style = ParagraphStyle(
        "BudgetSection",
        parent=styles["Heading2"],
        fontSize=10,
        leading=12,
        textColor=colors.HexColor(f"#{BLUE}"),
        spaceBefore=9,
        spaceAfter=4,
    )
    subsection_style = ParagraphStyle(
        "BudgetSubsection",
        parent=styles["Heading3"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#243B53"),
        spaceBefore=6,
        spaceAfter=3,
    )
    story = []
    school = resolve_tenant_school()
    if school:
        append_pdf_document_header(
            story,
            school,
            "COMPREHENSIVE BUDGET REPORT",
            show_statement_date=True,
            statement_date_text=f"Generated: {timezone.localtime(timezone.now()).strftime('%m/%d/%Y %I:%M %p')}",
            header_width_inches=10.3,
        )
    else:
        _, _, title_style, _ = get_pdf_header_styles()
        story.append(Paragraph("Comprehensive Budget Report", title_style))
    append_pdf_subtitle(
        story,
        f"{budget['name']} | {budget['academic_year']} | Report period {payload['context']['start_date']} to {payload['context']['end_date']} | All amounts in {currency}",
    )

    metadata = [
        ["Status", str(budget["status"]).title(), "Version", budget["version"], "Original Budget", "Yes" if budget["is_original"] else "No"],
        ["Academic Year", budget["academic_year"], "Year Dates", f"{budget['academic_year_start']} to {budget['academic_year_end']}", "Last Updated", _display_date(budget["updated_at"])],
        ["Submitted", _display_date(budget["submitted_at"]), "Approved", _display_date(budget["approved_at"]), "Activated", _display_date(budget["activated_at"])],
    ]
    meta_table = Table(metadata, colWidths=[0.9 * inch, 1.55 * inch] * 3)
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F7FA")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTNAME", (4, 0), (4, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B8C2CC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([meta_table, Paragraph("Planning Notes", subsection_style)])
    story.append(Paragraph(html.escape(budget["notes"] or "No planning notes provided."), body_style))

    summary = payload["summary"]
    revenue = details["revenue_composition"]
    story.append(Paragraph("Projected Revenue Composition", section_style))
    story.append(_pdf_table(
        ["Revenue Component", "Projected Amount"],
        [
            ["Revenue from Budget Lines", _amount(revenue["budget_line_revenue"])],
            ["Projected Student Tuition", _amount(revenue["projected_student_tuition"])],
            ["Projected Class-Section Fees", _amount(revenue["projected_class_section_fees"])],
            ["Total Projected Revenue", _amount(revenue["total_projected_revenue"])],
        ],
        body_style,
        widths=[3.2 * inch, 1.8 * inch],
    ))
    story.append(Paragraph("Financial Position", section_style))
    story.append(_pdf_table(
        ["Category", "Planned", "Actual", "Variance", "Performance"],
        [
            ["Total Revenue", _amount(summary["planned_revenue"]), _amount(summary["actual_revenue"]), _amount(summary["revenue_variance"]), format_percentage_display(summary["revenue_performance_percentage"])],
            ["Expense", _amount(summary["planned_expense"]), _amount(summary["actual_expense"]), _amount(summary["expense_variance"]), format_percentage_display(summary["expense_utilization_percentage"])],
            ["Surplus / (Deficit)", _amount(summary["planned_surplus"]), _amount(summary["actual_surplus"]), _amount(summary["surplus_variance"]), ""],
        ],
        body_style,
        widths=[2.1 * inch, 1.4 * inch, 1.4 * inch, 1.4 * inch, 1.4 * inch],
    ))

    enrollment_summary = details["enrollment_summary"]
    story.append(Paragraph("Enrollment and Student Fee Projection", section_style))
    story.append(Paragraph(
        f"<b>Academic Setup source:</b> {html.escape(details['definitions']['academic_setup'])}",
        body_style,
    ))
    story.append(Spacer(1, 4))
    story.append(_pdf_table(
        ["Prior Actual", "Budget Estimate", "Current Actual", "Projected Tuition", "Projected Other Fees", "Projected Student Fees"],
        [[
            enrollment_summary["prior_actual_students"],
            enrollment_summary["estimated_students"],
            enrollment_summary["actual_students"],
            _amount(enrollment_summary["projected_tuition"]),
            _amount(enrollment_summary["projected_other_fees"]),
            _amount(enrollment_summary["projected_student_fees"]),
        ]],
        body_style,
    ))
    if details["enrollment"]:
        story.append(Spacer(1, 4))
        story.append(_pdf_table(
            ["Grade", "Prior", "Estimate", "Actual", "Tuition / Student", "Other Fees / Student", "Total Fees / Student", "Projected Tuition", "Projected Total Fees", "Setup"],
            [[
                row["grade_level"], row["prior_actual_students"] or 0,
                row["estimated_students"], row["actual_students"],
                _amount(row["tuition_per_student"]), _amount(row["other_fees_per_student"]),
                _amount(row["total_fees_per_student"]), _amount(row["projected_tuition"]),
                _amount(row["projected_student_fees"]),
                "Ready" if row["setup_complete"] else "Review setup",
            ] for row in details["enrollment"]],
            body_style,
            widths=[1.05 * inch, 0.55 * inch, 0.65 * inch, 0.55 * inch, 1.05 * inch, 1.05 * inch, 1.05 * inch, 1.05 * inch, 1.1 * inch, 0.75 * inch],
            font_size=6,
        ))
        for row in details["enrollment"]:
            if row["setup_warnings"]:
                story.append(Paragraph(
                    f"<b>{html.escape(row['grade_level'])} setup:</b> {html.escape(' '.join(row['setup_warnings']))}",
                    body_style,
                ))

    workforce = details["workforce"]
    story.append(Paragraph("Workforce and Payroll Projection", section_style))
    story.append(_pdf_table(
        ["Compensation-Covered Employees", "Projected Base Payroll", "Actual Mapped Payroll (Report Period)", "Full-Year Variance"],
        [[workforce["compensation_covered_employees"], _amount(workforce["projected_base_payroll"]), _amount(workforce["actual_mapped_payroll_report_period"]), _amount(workforce["variance"])]],
        body_style,
    ))
    story.append(Paragraph(f"<b>Methodology:</b> {html.escape(workforce['methodology'])}", body_style))
    story.append(Paragraph(f"<b>Limitation:</b> {html.escape(workforce['limitation'])}", body_style))

    story.append(Paragraph("Budget Sections and Line Detail", section_style))
    for section in details["sections"]:
        totals = section["totals"]
        heading = f"{section['name']} ({section['section_type'].title()}) - {section['line_count']} line(s)"
        section_content = [Paragraph(html.escape(heading), subsection_style)]
        if section["lines"]:
            section_content.append(_pdf_table(
                ["Budget Line", "GL Account", "Source", "Original", "Amendments", "Effective", "Report Plan", "Actual", "Variance"],
                [[
                    line["name"],
                    " - ".join(value for value in (line["gl_account_code"], line["gl_account_name"]) if value),
                    line["source_type"].replace("_", " ").title(),
                    _amount(line["original_amount"]), _amount(line["approved_amendments"]),
                    _amount(line["effective_amount"]), _amount(line["report_planned_amount"]),
                    _amount(line["actual_amount"]), _amount(line["variance_amount"]),
                ] for line in section["lines"]],
                body_style,
                widths=[1.4 * inch, 1.45 * inch, 0.75 * inch, 0.85 * inch, 0.85 * inch, 0.85 * inch, 0.85 * inch, 0.85 * inch, 0.85 * inch],
                font_size=6,
            ))
            section_content.append(_pdf_table(
                ["Section Total", "Original", "Amendments", "Effective", "Report Plan", "Actual", "Variance"],
                [[
                    section["name"], _amount(totals.get("original_amount")),
                    _amount(totals.get("approved_amendments")), _amount(totals.get("effective_amount")),
                    _amount(totals.get("report_planned_amount")), _amount(totals.get("actual_amount")),
                    _amount(totals.get("variance_amount")),
                ]],
                body_style,
                font_size=6,
            ))
        else:
            section_content.append(Paragraph("No plan lines in this section.", body_style))
        story.extend(section_content)

    story.append(Paragraph("Period Allocations", section_style))
    if details["period_allocations"]:
        story.append(_pdf_table(
            ["Section", "Budget Line", "Start Date", "End Date", "Planned Amount"],
            [[row["section"], row["line"], row["start_date"], row["end_date"], _amount(row["planned_amount"])] for row in details["period_allocations"]],
            body_style,
            widths=[1.8 * inch, 2.6 * inch, 1.3 * inch, 1.3 * inch, 1.4 * inch],
        ))
    else:
        story.append(Paragraph("No period allocations recorded.", body_style))

    story.append(Paragraph("Budget Amendments", section_style))
    if details["revisions"]:
        for revision in details["revisions"]:
            story.append(Paragraph(
                f"Amendment {revision['number']} - {str(revision['status']).title()}: {html.escape(revision['reason'])}",
                subsection_style,
            ))
            if revision["line_deltas"]:
                story.append(_pdf_table(
                    ["Section", "Budget Line", "Amount Change", "Rationale"],
                    [[row["section"], row["line"], _amount(row["amount_delta"]), row["rationale"]] for row in revision["line_deltas"]],
                    body_style,
                    widths=[1.7 * inch, 2.2 * inch, 1.4 * inch, 4.2 * inch],
                ))
    else:
        story.append(Paragraph("No budget amendments recorded.", body_style))

    story.append(Paragraph("Lifecycle History", section_style))
    if details["lifecycle"]:
        story.append(_pdf_table(
            ["Date", "Event", "From", "To", "Actor", "Reason"],
            [[_display_date(row["created_at"]), row["event_type"].replace("_", " ").title(), row["from_status"].title(), row["to_status"].title(), row["actor"], row["reason"]] for row in details["lifecycle"]],
            body_style,
            widths=[1.2 * inch, 1.5 * inch, 0.9 * inch, 0.9 * inch, 1.6 * inch, 3.5 * inch],
        ))
    else:
        story.append(Paragraph("No lifecycle events recorded.", body_style))

    story.append(Paragraph("Definitions and Methodology", section_style))
    for key, value in details["definitions"].items():
        story.append(Paragraph(f"<b>{html.escape(key.replace('_', ' ').title())}:</b> {html.escape(str(value))}", body_style))

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _write_sheet_table(ws, headers, rows, currency_code, *, start_row=1):
    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            header = headers[col_idx - 1]
            prepared = prepare_xlsx_numeric_value(header, value)
            cell = ws.cell(row=row_idx, column=col_idx, value=prepared)
            apply_xlsx_cell_style(cell, header, prepared, currency_code)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        values = [str(header)] + [str(row[col_idx - 1] or "") for row in rows]
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(map(len, values)) + 2, 42)
    ws.freeze_panes = f"A{start_row + 1}"


def build_comprehensive_budget_xlsx(payload, filename):
    details = payload["details"]
    budget = details["budget"]
    currency = budget["currency"]
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    overview["A1"] = "Comprehensive Budget Report"
    overview["A1"].font = Font(bold=True, size=16, color=BLUE)
    overview["A2"] = f"{budget['name']} | {budget['academic_year']} | All amounts in {currency}"
    overview_rows = [
        ("Budget Name", budget["name"]), ("Status", str(budget["status"]).title()),
        ("Version", budget["version"]), ("Original Budget", "Yes" if budget["is_original"] else "No"),
        ("Academic Year", budget["academic_year"]),
        ("Academic Year Dates", f"{budget['academic_year_start']} to {budget['academic_year_end']}"),
        ("Report Period", f"{payload['context']['start_date']} to {payload['context']['end_date']}"),
        ("Currency", currency), ("Planning Notes", budget["notes"] or "No planning notes provided."),
        ("Submitted", _display_date(budget["submitted_at"])), ("Submitted By", budget["submitted_by"]),
        ("Approved", _display_date(budget["approved_at"])), ("Approved By", budget["approved_by"]),
        ("Activated", _display_date(budget["activated_at"])), ("Activated By", budget["activated_by"]),
        ("Closed", _display_date(budget["closed_at"])), ("Closed By", budget["closed_by"]),
        ("Last Updated", _display_date(budget["updated_at"])),
    ]
    for row_idx, (label, value) in enumerate(overview_rows, 4):
        overview.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        overview.cell(row=row_idx, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
    revenue_row = len(overview_rows) + 6
    overview.cell(row=revenue_row - 1, column=1, value="Projected Revenue Composition").font = Font(bold=True, size=12, color=BLUE)
    summary = payload["summary"]
    revenue = details["revenue_composition"]
    _write_sheet_table(
        overview,
        ["Revenue Component", "Projected Amount"],
        [
            ["Revenue from Budget Lines", revenue["budget_line_revenue"]],
            ["Projected Student Tuition", revenue["projected_student_tuition"]],
            ["Projected Class-Section Fees", revenue["projected_class_section_fees"]],
            ["Total Projected Revenue", revenue["total_projected_revenue"]],
        ],
        currency,
        start_row=revenue_row,
    )
    financial_row = revenue_row + 7
    overview.cell(row=financial_row - 1, column=1, value="Financial Position").font = Font(bold=True, size=12, color=BLUE)
    _write_sheet_table(
        overview,
        ["Category", "Planned Amount", "Actual Amount", "Variance Amount", "Performance Percentage"],
        [
            ["Total Revenue", summary["planned_revenue"], summary["actual_revenue"], summary["revenue_variance"], summary["revenue_performance_percentage"]],
            ["Expense", summary["planned_expense"], summary["actual_expense"], summary["expense_variance"], summary["expense_utilization_percentage"]],
            ["Surplus / (Deficit)", summary["planned_surplus"], summary["actual_surplus"], summary["surplus_variance"], None],
        ],
        currency,
        start_row=financial_row,
    )

    plan_rows = []
    for section in details["sections"]:
        for line in section["lines"]:
            plan_rows.append([
                section["name"], section["section_type"].title(), line["name"], line["gl_account_code"],
                line["gl_account_name"], line["source_type"].replace("_", " ").title(), line["source_ref"],
                line["original_amount"], line["approved_amendments"], line["effective_amount"],
                line["report_planned_amount"], line["actual_amount"], line["variance_amount"],
                line["variance_percentage"],
            ])
    plan = wb.create_sheet("Plan Details")
    _write_sheet_table(
        plan,
        ["Section", "Type", "Budget Line", "GL Code", "GL Account", "Source", "Source Reference", "Original Amount", "Approved Amendments", "Effective Amount", "Report Planned Amount", "Actual Amount", "Variance Amount", "Variance Percentage"],
        plan_rows,
        currency,
    )

    periods = wb.create_sheet("Period Allocations")
    _write_sheet_table(
        periods,
        ["Section", "Budget Line", "Start Date", "End Date", "Planned Amount"],
        [[row["section"], row["line"], row["start_date"], row["end_date"], row["planned_amount"]] for row in details["period_allocations"]],
        currency,
    )

    enrollment = wb.create_sheet("Enrollment")
    _write_sheet_table(
        enrollment,
        ["Grade Level", "Student Category", "Prior Actual Students", "Estimated Students", "Actual Students", "Headcount Variance", "Tuition Per Student", "Other Fees Per Student", "Total Fees Per Student", "Projected Tuition", "Projected Other Fees", "Projected Student Fees", "Active Sections", "Setup Status", "Setup Warnings"],
        [[row["grade_level"], row["student_category"] or "Returning", row["prior_actual_students"] or 0, row["estimated_students"], row["actual_students"], row["headcount_variance"], row["tuition_per_student"], row["other_fees_per_student"], row["total_fees_per_student"], row["projected_tuition"], row["projected_other_fees"], row["projected_student_fees"], row["section_count"], "Ready" if row["setup_complete"] else "Review setup", " ".join(row["setup_warnings"])] for row in details["enrollment"]],
        currency,
    )

    workforce = wb.create_sheet("Workforce")
    staffing = details["workforce"]
    _write_sheet_table(
        workforce,
        ["Compensation-Covered Employees", "Projected Base Payroll", "Actual Mapped Payroll (Report Period)", "Full-Year Variance Amount", "Staffing Plan Available"],
        [[staffing["compensation_covered_employees"], staffing["projected_base_payroll"], staffing["actual_mapped_payroll_report_period"], staffing["variance"], "Yes" if staffing["staffing_plan_available"] else "No"]],
        currency,
    )
    workforce["A4"] = "Methodology"
    workforce["B4"] = staffing["methodology"]
    workforce["A5"] = "Limitation"
    workforce["B5"] = staffing["limitation"]
    workforce.column_dimensions["B"].width = 80
    workforce["B4"].alignment = Alignment(wrap_text=True)
    workforce["B5"].alignment = Alignment(wrap_text=True)

    amendments = wb.create_sheet("Amendments")
    amendment_rows = []
    for revision in details["revisions"]:
        if revision["line_deltas"]:
            for delta in revision["line_deltas"]:
                amendment_rows.append([revision["number"], str(revision["status"]).title(), revision["reason"], _display_date(revision["approved_at"]), revision["approved_by"], delta["section"], delta["line"], delta["amount_delta"], delta["rationale"]])
        else:
            amendment_rows.append([revision["number"], str(revision["status"]).title(), revision["reason"], _display_date(revision["approved_at"]), revision["approved_by"], "", "", None, ""])
    _write_sheet_table(
        amendments,
        ["Amendment", "Status", "Reason", "Approved Date", "Approved By", "Section", "Budget Line", "Amount Change", "Rationale"],
        amendment_rows,
        currency,
    )

    lifecycle = wb.create_sheet("Lifecycle")
    _write_sheet_table(
        lifecycle,
        ["Date", "Event", "From Status", "To Status", "Actor", "Reason"],
        [[_display_date(row["created_at"]), row["event_type"].replace("_", " ").title(), row["from_status"].title(), row["to_status"].title(), row["actor"], row["reason"]] for row in details["lifecycle"]],
        currency,
    )

    definitions = wb.create_sheet("Definitions")
    _write_sheet_table(
        definitions,
        ["Term", "Definition"],
        [[key.replace("_", " ").title(), value] for key, value in details["definitions"].items()],
        currency,
    )
    definitions.column_dimensions["B"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_comprehensive_budget_report(request, payload, filename_base):
    export_format = get_export_format(request)
    if export_format == "pdf":
        return build_comprehensive_budget_pdf(request, payload, f"{filename_base}.pdf")
    if export_format == "xlsx":
        return build_comprehensive_budget_xlsx(payload, f"{filename_base}.xlsx")
    return None
