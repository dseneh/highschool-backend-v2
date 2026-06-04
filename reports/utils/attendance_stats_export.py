"""Excel and PDF export layout for daily attendance statistics."""

from __future__ import annotations

import io
from datetime import datetime

from django.http import HttpResponse

GROUP_LABELS = ("Total Students", "Present", "Tardy", "Absent")
SUB_LABELS = ("Males", "Females", "Total")
LAST_COL = 13


def _whole_int(value) -> int:
    try:
        return int(round(float(value or 0)))
    except (TypeError, ValueError):
        return 0


def _whole_pct(value) -> int:
    return _whole_int(value)


def _counts_triplet(block: dict | None) -> tuple[int, int, int]:
    block = block or {}
    return (
        _whole_int(block.get("male")),
        _whole_int(block.get("female")),
        _whole_int(block.get("total")),
    )


def _pct_triplet_from_percentages(percentages: dict, key: str) -> tuple[int, int, int]:
    block = percentages.get(key) or {}
    return (
        _whole_pct(block.get("male")),
        _whole_pct(block.get("female")),
        _whole_pct(block.get("total")),
    )


def _section_row(section: dict) -> list[int | str]:
    return [
        section.get("class_label") or section.get("section_name") or "",
        *_counts_triplet(section.get("total_students")),
        *_counts_triplet(section.get("present")),
        *_counts_triplet(section.get("tardy")),
        *_counts_triplet(section.get("absent")),
    ]


def _format_report_title(target_date: str) -> str:
    try:
        parsed = datetime.strptime(target_date, "%Y-%m-%d")
        return f"Attendance Stats for {parsed.strftime('%B %d, %Y')}"
    except ValueError:
        return f"Attendance Stats for {target_date}"


def _excel_styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    return {
        "border": border,
        "header_fill": header_fill,
        "header_font": header_font,
        "title_font": title_font,
        "label_font": label_font,
        "center": center,
        "left": left,
    }


def _apply_border_cell(ws, row: int, col: int, value, *, styles, align_center=True, bold=False):
    from openpyxl.styles import Font

    cell = ws.cell(row=row, column=col, value=value)
    cell.border = styles["border"]
    cell.alignment = styles["center"] if align_center else styles["left"]
    if bold:
        cell.font = styles["label_font"]
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.font = Font(size=10)
    else:
        cell.font = Font(size=10)
    return cell


def _write_pct_group_header_row(ws, row: int, *, styles) -> None:
    _apply_border_cell(ws, row, 1, "", styles=styles)
    for group_idx, label in enumerate(("Present", "Tardy", "Absent")):
        start_col = 2 + group_idx * 3
        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=start_col + 2,
        )
        cell = ws.cell(row=row, column=start_col, value=label)
        cell.border = styles["border"]
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        for offset in (1, 2):
            extra = ws.cell(row=row, column=start_col + offset)
            extra.border = styles["border"]
            extra.fill = styles["header_fill"]


def _write_group_header_row(ws, row: int, *, styles, class_header: str = "Class") -> None:
    _apply_border_cell(ws, row, 1, class_header, styles=styles, align_center=False, bold=True)
    for group_idx, label in enumerate(GROUP_LABELS):
        start_col = 2 + group_idx * 3
        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=start_col + 2,
        )
        cell = ws.cell(row=row, column=start_col, value=label)
        cell.border = styles["border"]
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        for offset in (1, 2):
            extra = ws.cell(row=row, column=start_col + offset)
            extra.border = styles["border"]
            extra.fill = styles["header_fill"]


def _write_subheader_row(ws, row: int, *, styles, first_col_blank=False) -> None:
    if first_col_blank:
        _apply_border_cell(ws, row, 1, "", styles=styles)
    else:
        _apply_border_cell(ws, row, 1, "", styles=styles, align_center=False)
    for group_idx in range(4):
        for sub_idx, sub in enumerate(SUB_LABELS):
            col = 2 + group_idx * 3 + sub_idx
            cell = ws.cell(row=row, column=col, value=sub)
            cell.border = styles["border"]
            cell.font = styles["header_font"]
            cell.alignment = styles["center"]


def build_attendance_stats_xlsx(*, payload: dict, filename: str) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    styles = _excel_styles()
    target_date = payload.get("date", "")
    percentages = payload.get("percentages") or {}
    sections = payload.get("sections") or []
    totals = payload.get("totals") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Stats"

    title = _format_report_title(target_date)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = styles["title_font"]
    title_cell.alignment = styles["center"]

    row = 3
    _write_pct_group_header_row(ws, row, styles=styles)
    row += 1
    for group_idx in range(3):
        for sub_idx, sub in enumerate(SUB_LABELS):
            col = 2 + group_idx * 3 + sub_idx
            cell = ws.cell(row=row, column=col, value=sub)
            cell.border = styles["border"]
            cell.font = styles["header_font"]
            cell.alignment = styles["center"]
    _apply_border_cell(ws, row, 1, "", styles=styles)
    row += 1

    _apply_border_cell(
        ws,
        row,
        1,
        "Percentages - All Classes:",
        styles=styles,
        align_center=False,
        bold=True,
    )
    pct_values = []
    for key in ("present", "tardy", "absent"):
        pct_values.extend(_pct_triplet_from_percentages(percentages, key))
    for col_offset, pct in enumerate(pct_values, start=2):
        _apply_border_cell(ws, row, col_offset, pct, styles=styles)
    row += 2

    _write_group_header_row(ws, row, styles=styles)
    row += 1
    _write_subheader_row(ws, row, styles=styles)
    row += 1

    for section in sections:
        values = _section_row(section)
        for col_idx, value in enumerate(values, start=1):
            _apply_border_cell(
                ws,
                row,
                col_idx,
                value,
                styles=styles,
                align_center=col_idx > 1,
                bold=col_idx == 1,
            )
        row += 1

    if totals:
        totals_values = [
            "Totals",
            *_counts_triplet(totals.get("total_students")),
            *_counts_triplet(totals.get("present")),
            *_counts_triplet(totals.get("tardy")),
            *_counts_triplet(totals.get("absent")),
        ]
        for col_idx, value in enumerate(totals_values, start=1):
            _apply_border_cell(
                ws,
                row,
                col_idx,
                value,
                styles=styles,
                align_center=col_idx > 1,
                bold=True,
            )

    ws.column_dimensions["A"].width = 28
    for col in range(2, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_attendance_stats_pdf(*, request, payload: dict, filename: str) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from common.services.pdf_components import (
        append_pdf_document_header,
        resolve_tenant_school,
    )

    styles = _excel_styles()
    target_date = payload.get("date", "")
    percentages = payload.get("percentages") or {}
    sections = payload.get("sections") or []
    totals = payload.get("totals") or {}

    school = resolve_tenant_school(None)
    title = _format_report_title(target_date)

    buf = io.BytesIO()
    page_size = landscape(letter)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    story: list = []

    if school:
        append_pdf_document_header(
            story,
            school,
            title.upper(),
            show_statement_date=False,
            bottom_spacer_inches=0.08,
        )
    else:
        story.append(Paragraph(title, getSampleStyleSheet()["Title"]))

    story.append(Spacer(1, 10))

    grid_border = 0.5
    header_bg = colors.HexColor("#D9D9D9")

    def _table_style(col_count: int, header_rows: int) -> TableStyle:
        return TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), grid_border, colors.black),
                ("FONTNAME", (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (0, header_rows - 1), header_bg),
                ("BACKGROUND", (1, 0), (col_count - 1, header_rows - 1), header_bg),
            ]
        )

    from reportlab.lib.styles import getSampleStyleSheet

    pct_col_widths = [1.6 * inch] + [0.55 * inch] * 9
    pct_header_row1 = [""]
    for label in ("Present", "Tardy", "Absent"):
        pct_header_row1.extend([label, "", ""])

    pct_header_row2 = [""]
    for _ in range(3):
        pct_header_row2.extend(SUB_LABELS)

    pct_data_row = ["Percentages - All Classes:"]
    for key in ("present", "tardy", "absent"):
        male, female, total = _pct_triplet_from_percentages(percentages, key)
        pct_data_row.extend([str(male), str(female), str(total)])

    pct_table = Table(
        [pct_header_row1, pct_header_row2, pct_data_row],
        colWidths=pct_col_widths,
    )
    pct_style = _table_style(10, 2)
    pct_style.add("SPAN", (1, 0), (3, 0))
    pct_style.add("SPAN", (4, 0), (6, 0))
    pct_style.add("SPAN", (7, 0), (9, 0))
    pct_table.setStyle(pct_style)
    story.append(pct_table)
    story.append(Spacer(1, 14))

    class_header_row1 = ["Class"]
    for label in GROUP_LABELS:
        class_header_row1.extend([label, "", ""])
    class_header_row2 = [""]
    for _ in range(4):
        class_header_row2.extend(SUB_LABELS)

    class_rows: list[list[str]] = [class_header_row1, class_header_row2]
    for section in sections:
        row = _section_row(section)
        class_rows.append([str(v) for v in row])
    if totals:
        class_rows.append(
            [
                str(v)
                for v in [
                    "Totals",
                    *_counts_triplet(totals.get("total_students")),
                    *_counts_triplet(totals.get("present")),
                    *_counts_triplet(totals.get("tardy")),
                    *_counts_triplet(totals.get("absent")),
                ]
            ]
        )

    class_table = Table(class_rows, colWidths=[1.6 * inch] + [0.55 * inch] * 12, repeatRows=2)
    class_style = _table_style(LAST_COL, 2)
    for group_idx in range(4):
        start = 1 + group_idx * 3
        class_style.add("SPAN", (0, 0), (0, 1))
        class_style.add("SPAN", (start, 0), (start + 2, 0))
    if totals:
        class_style.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    class_table.setStyle(class_style)
    story.append(class_table)

    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
