"""Shared helpers for report views."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response


def parse_decimal_param(value: str | None) -> Decimal | None:
    if value is None:
        return None
    normalized = str(value).strip()
    if not normalized:
        return None
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None


def read_multi_query_values(request, key: str) -> list[str]:
    values: list[str] = []
    for raw in request.query_params.getlist(key):
        if raw is None:
            continue
        for part in str(raw).split(","):
            value = part.strip()
            if value:
                values.append(value)
    if not values:
        single = request.query_params.get(key)
        if single:
            for part in str(single).split(","):
                value = part.strip()
                if value:
                    values.append(value)
    return list(dict.fromkeys(values))


def resolve_academic_year(request):
    from academics.models import AcademicYear

    academic_year_id = request.query_params.get("academic_year_id")
    if academic_year_id:
        try:
            return AcademicYear.objects.get(id=academic_year_id), None
        except AcademicYear.DoesNotExist:
            return None, Response({"detail": "Academic year not found."}, status=status.HTTP_404_NOT_FOUND)

    academic_year = AcademicYear.objects.filter(current=True).first()
    if not academic_year:
        return None, Response({"detail": "No current academic year found."}, status=status.HTTP_400_BAD_REQUEST)
    return academic_year, None


def parse_date_param(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def wants_xlsx(request) -> bool:
    return get_export_format(request) == "xlsx"


def wants_pdf(request) -> bool:
    return get_export_format(request) == "pdf"


def get_export_format(request) -> str | None:
    fmt = request.query_params.get("export") or request.query_params.get("format")
    normalized = str(fmt or "").strip().lower()
    if normalized in {"xlsx", "excel"}:
        return "xlsx"
    if normalized == "pdf":
        return "pdf"
    return None


def build_pdf_response(
    *,
    request=None,
    school=None,
    filename: str,
    title: str,
    subtitle: str | None,
    headers: list[str],
    rows: list[list[object]],
    show_generated_date: bool = True,
) -> HttpResponse:
    from django.utils import timezone
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from common.services.pdf_components import (
        append_pdf_document_header,
        append_pdf_subtitle,
        get_pdf_header_styles,
        resolve_tenant_school,
    )

    school = resolve_tenant_school(school)

    buf = io.BytesIO()
    page_size = landscape(letter)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.3 * inch,
        bottomMargin=0.3 * inch,
    )
    content_width_inches = (page_size[0] - doc.leftMargin - doc.rightMargin) / inch

    story: list = []
    if school:
        statement_date_text = ""
        if show_generated_date:
            statement_date_text = (
                f"Generated: {timezone.localtime(timezone.now()).strftime('%m/%d/%Y %I:%M %p')}"
            )
        append_pdf_document_header(
            story,
            school,
            title.upper(),
            show_statement_date=bool(statement_date_text),
            statement_date_text=statement_date_text,
            bottom_spacer_inches=0.04,
            header_width_inches=content_width_inches,
        )
    else:
        _, _, title_style, _ = get_pdf_header_styles()
        story.append(Paragraph(title, title_style))

    if subtitle:
        append_pdf_subtitle(story, subtitle)
    story.append(Spacer(1, 8))

    table_data = [headers] + [[str(cell) if cell is not None else "" for cell in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_xlsx_response(
    *,
    filename: str,
    title: str,
    subtitle: str | None,
    summary_rows: list[tuple[str, object]] | None,
    headers: list[str],
    rows: list[list[object]],
    column_widths: list[int] | None = None,
) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    row_idx = 2
    if subtitle:
        ws[f"A{row_idx}"] = subtitle
        row_idx += 1

    if summary_rows:
        row_idx += 1
        for label, value in summary_rows:
            ws[f"A{row_idx}"] = label
            ws[f"B{row_idx}"] = value
            row_idx += 1
        row_idx += 1

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = row_idx
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for offset, row in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=offset, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = Font(size=9)
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = "#,##0.00"

    widths = column_widths or [max(len(str(h)), 12) for h in headers]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_tabular_report(
    request,
    *,
    filename_base: str,
    title: str,
    subtitle: str | None,
    summary_rows: list[tuple[str, object]] | None,
    headers: list[str],
    rows: list[list[object]],
    column_widths: list[int] | None = None,
):
    export_format = get_export_format(request)
    if export_format == "pdf":
        return build_pdf_response(
            request=request,
            filename=f"{filename_base}.pdf",
            title=title,
            subtitle=subtitle,
            headers=headers,
            rows=rows,
        )
    if export_format == "xlsx":
        return build_xlsx_response(
            filename=f"{filename_base}.xlsx",
            title=title,
            subtitle=subtitle,
            summary_rows=summary_rows,
            headers=headers,
            rows=rows,
            column_widths=column_widths,
        )
    return None
