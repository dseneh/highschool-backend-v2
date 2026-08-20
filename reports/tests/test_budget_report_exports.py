from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch
from uuid import uuid4

from django.http import HttpResponse
from django.test import SimpleTestCase
from openpyxl import load_workbook
from pypdf import PdfReader

from reports.views.budget_reports import BudgetReportView


class BudgetReportExportParityTests(SimpleTestCase):
    def setUp(self):
        self.budget_id = uuid4()
        self.budget = SimpleNamespace(
            id=self.budget_id,
            academic_year=SimpleNamespace(
                start_date=date(2025, 9, 1), end_date=date(2026, 6, 30)
            ),
        )
        self.payload = {
            "summary": {
                "planned_revenue": Decimal("210.00"),
                "actual_revenue": Decimal("90.00"),
                "revenue_variance": Decimal("-120.00"),
                "revenue_performance_percentage": Decimal("42.86"),
                "planned_expense": Decimal("60.00"),
                "actual_expense": Decimal("55.00"),
                "expense_variance": Decimal("5.00"),
                "expense_utilization_percentage": Decimal("91.67"),
                "planned_surplus": Decimal("150.00"),
                "actual_surplus": Decimal("35.00"),
                "surplus_variance": Decimal("-115.00"),
            },
            "results": [{"line": "Tuition", "planned_amount": Decimal("100.00")}],
            "count": 1,
            "context": {
                "academic_year": "2025/26",
                "start_date": "2025-09-01",
                "end_date": "2026-06-30",
            },
            "definitions": {"planned": "Approved budget plan."},
            "columns": [["Budget Line", "line"], ["Planned Amount", "planned_amount"]],
            "details": {
                "budget": {
                    "name": "Annual Budget",
                    "status": "approved",
                    "version": 2,
                    "is_original": True,
                    "notes": "Fund the annual school operating plan.",
                    "academic_year": "2025/26",
                    "academic_year_start": "2025-09-01",
                    "academic_year_end": "2026-06-30",
                    "currency": "USD",
                    "submitted_at": None,
                    "submitted_by": "",
                    "approved_at": None,
                    "approved_by": "",
                    "activated_at": None,
                    "activated_by": "",
                    "closed_at": None,
                    "closed_by": "",
                    "created_at": "2025-08-01T00:00:00+00:00",
                    "updated_at": "2025-08-02T00:00:00+00:00",
                },
                "sections": [{
                    "id": "section-1",
                    "name": "Revenue",
                    "section_type": "revenue",
                    "line_count": 1,
                    "totals": {
                        "original_amount": Decimal("100.00"),
                        "approved_amendments": Decimal("0.00"),
                        "effective_amount": Decimal("100.00"),
                        "report_planned_amount": Decimal("100.00"),
                        "actual_amount": Decimal("90.00"),
                        "variance_amount": Decimal("-10.00"),
                    },
                    "lines": [{
                        "id": "line-1",
                        "name": "Tuition",
                        "source_type": "fee_rate",
                        "source_ref": "",
                        "gl_account_code": "4000",
                        "gl_account_name": "Tuition Revenue",
                        "original_amount": Decimal("100.00"),
                        "approved_amendments": Decimal("0.00"),
                        "effective_amount": Decimal("100.00"),
                        "report_planned_amount": Decimal("100.00"),
                        "actual_amount": Decimal("90.00"),
                        "variance_amount": Decimal("-10.00"),
                        "variance_percentage": Decimal("-10.00"),
                        "periods": [],
                    }],
                }],
                "period_allocations": [],
                "enrollment": [{
                    "grade_level": "Grade 1",
                    "student_category": "",
                    "prior_actual_students": 20,
                    "estimated_students": 24,
                    "actual_students": 22,
                    "headcount_variance": -2,
                    "tuition_per_student": Decimal("100.00"),
                    "other_fees_per_student": Decimal("10.00"),
                    "total_fees_per_student": Decimal("110.00"),
                    "projected_tuition": Decimal("100.00"),
                    "projected_other_fees": Decimal("10.00"),
                    "projected_student_fees": Decimal("110.00"),
                    "section_count": 2,
                    "setup_complete": True,
                    "setup_warnings": [],
                }],
                "enrollment_summary": {
                    "prior_actual_students": 20,
                    "estimated_students": 24,
                    "actual_students": 22,
                    "projected_tuition": Decimal("100.00"),
                    "projected_other_fees": Decimal("10.00"),
                    "projected_student_fees": Decimal("110.00"),
                },
                "revenue_composition": {
                    "budget_line_revenue": Decimal("100.00"),
                    "projected_student_tuition": Decimal("100.00"),
                    "projected_class_section_fees": Decimal("10.00"),
                    "projected_student_fees": Decimal("110.00"),
                    "total_projected_revenue": Decimal("210.00"),
                },
                "workforce": {
                    "compensation_covered_employees": 12,
                    "projected_base_payroll": Decimal("60.00"),
                    "actual_mapped_payroll_report_period": Decimal("55.00"),
                    "variance": Decimal("5.00"),
                    "staffing_plan_available": False,
                    "methodology": "Effective-dated compensation records.",
                    "limitation": "This is not an approved staffing or FTE plan.",
                },
                "revisions": [],
                "lifecycle": [],
                "definitions": {
                    "planned": "Approved budget plan.",
                    "academic_setup": "Configure tuition and section fees before budgeting.",
                },
            },
        }

    def _request(self, export=None):
        params = {"budget_id": str(self.budget_id)}
        if export:
            params["export"] = export
        return SimpleNamespace(query_params=params)

    @patch("reports.views.budget_reports.build_budget_report")
    @patch("reports.views.budget_reports.Budget.objects.select_related")
    def test_json_returns_canonical_payload(self, mock_select, mock_build):
        mock_select.return_value.prefetch_related.return_value.get.return_value = self.budget
        mock_build.return_value = self.payload
        view = BudgetReportView()
        view.report_type = "budget-summary"

        response = view.get(self._request())

        self.assertEqual(response.data, self.payload)

    @patch("reports.views.budget_reports.export_tabular_report")
    @patch("reports.views.budget_reports.build_budget_report")
    @patch("reports.views.budget_reports.Budget.objects.select_related")
    def test_pdf_and_xlsx_receive_identical_canonical_rows(
        self, mock_select, mock_build, mock_export
    ):
        mock_select.return_value.prefetch_related.return_value.get.return_value = self.budget
        mock_build.return_value = self.payload
        mock_export.side_effect = [HttpResponse(b"pdf"), HttpResponse(b"xlsx")]
        view = BudgetReportView()
        view.report_type = "budget-vs-actual"

        view.get(self._request("pdf"))
        view.get(self._request("xlsx"))

        first = mock_export.call_args_list[0].kwargs
        second = mock_export.call_args_list[1].kwargs
        expected_rows = [["Tuition", Decimal("100.00")]]
        self.assertEqual(first["rows"], expected_rows)
        self.assertEqual(second["rows"], expected_rows)
        self.assertEqual(first["summary_rows"], second["summary_rows"])

    @patch("reports.views.budget_reports.build_budget_report")
    @patch("reports.views.budget_reports.Budget.objects.select_related")
    def test_real_export_renderers_return_pdf_and_excel_files(self, mock_select, mock_build):
        mock_select.return_value.prefetch_related.return_value.get.return_value = self.budget
        mock_build.return_value = self.payload
        view = BudgetReportView()
        view.report_type = "budget-summary"

        pdf = view.get(self._request("pdf"))
        xlsx = view.get(self._request("xlsx"))

        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))
        self.assertEqual(
            xlsx["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(xlsx.content.startswith(b"PK"))

        pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages)
        self.assertIn("Projected Revenue Composition", pdf_text)
        self.assertIn("Projected Student Tuition", pdf_text)
        self.assertIn("Projected Class-Section Fees", pdf_text)
        self.assertIn("Total Projected Revenue", pdf_text)
        self.assertIn("Enrollment and Student Fee Projection", pdf_text)
        self.assertIn("Workforce and Payroll Projection", pdf_text)
        self.assertIn("Budget Sections and Line Detail", pdf_text)

        workbook = load_workbook(BytesIO(xlsx.content), data_only=True)
        self.assertEqual(
            set(workbook.sheetnames),
            {
                "Overview", "Plan Details", "Period Allocations", "Enrollment",
                "Workforce", "Amendments", "Lifecycle", "Definitions",
            },
        )
        self.assertEqual(workbook["Enrollment"]["A2"].value, "Grade 1")
        self.assertEqual(workbook["Workforce"]["A2"].value, 12)
