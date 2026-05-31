"""Combined income and expense summary reports from accounting cash transactions."""

from __future__ import annotations

import io

from django.db.models import Count, F, Q, Sum
from django.db.models.functions import Abs, TruncMonth
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework.views import APIView

from accounting.models import AccountingCashTransaction
from accounting.services.post_all import apply_cash_transaction_list_filters

from ..accounting_totals import (
    EXPENSE_CATEGORY,
    INCOME_CATEGORY,
    approved_cash_queryset,
    filter_cash_by_period,
    sum_approved_cash_net,
    sum_period_income_minus_expense,
)
from ..access_policies import ReportsAccessPolicy
from accounting.services.currency_totals import get_tenant_base_currency, serialize_currency

INCOME = INCOME_CATEGORY
EXPENSE = EXPENSE_CATEGORY
LEDGER_CATEGORIES = Q(transaction_type__transaction_category__in=("income", "expense"))

GROUP_BY_OPTIONS = {
    "bank_account": {
        "values": (
            "bank_account_id",
            "bank_account__account_name",
            "bank_account__bank_name",
            "bank_account__account_number",
        ),
        "label_fields": ("bank_account__account_name", "bank_account__bank_name"),
    },
    "transaction_type": {
        "values": (
            "transaction_type_id",
            "transaction_type__code",
            "transaction_type__name",
            "transaction_type__transaction_category",
        ),
        "label_fields": ("transaction_type__name", "transaction_type__code"),
    },
    "payment_method": {
        "values": (
            "payment_method_id",
            "payment_method__name",
            "payment_method__code",
        ),
        "label_fields": ("payment_method__name", "payment_method__code"),
    },
    "month": {
        "values": ("month",),
        "label_fields": ("month",),
        "annotate_month": True,
    },
}


class AccountingSummaryReportView(APIView):
    """Summarise income and expense cash transactions with flexible grouping."""

    permission_classes = [ReportsAccessPolicy]

    @staticmethod
    def _parse_group_by(value: str | None) -> str | None:
        normalized = (value or "all").strip().lower()
        if normalized in {"", "all"}:
            return None
        return normalized if normalized in GROUP_BY_OPTIONS else "bank_account"

    def _compute_groups(self, queryset, group_by: str) -> list[dict]:
        group_config = GROUP_BY_OPTIONS[group_by]
        grouped_queryset = queryset
        if group_config.get("annotate_month"):
            grouped_queryset = grouped_queryset.annotate(month=TruncMonth("transaction_date"))

        grouped_rows = (
            grouped_queryset.values(*group_config["values"])
            .annotate(
                income=Sum(Abs(F("base_amount")), filter=INCOME),
                expense=Sum(Abs(F("base_amount")), filter=EXPENSE),
                income_count=Count("id", filter=INCOME),
                expense_count=Count("id", filter=EXPENSE),
            )
            .order_by("-income", "-expense")
        )

        return [self._build_group_row(row, group_by) for row in grouped_rows]

    @staticmethod
    def _build_group_row(row: dict, group_by: str) -> dict:
        bank_name = ""
        account_number = ""

        if group_by == "bank_account":
            group_id = str(row.get("bank_account_id") or "")
            label = row.get("bank_account__account_name") or "Unknown account"
            bank_name = (row.get("bank_account__bank_name") or "").strip()
            account_number = (row.get("bank_account__account_number") or "").strip()
            detail_parts = [part for part in (bank_name, account_number) if part]
            sub_label = " · ".join(detail_parts)
        elif group_by == "transaction_type":
            group_id = str(row.get("transaction_type_id") or "")
            label = row.get("transaction_type__name") or row.get("transaction_type__code") or "Unknown type"
            sub_label = row.get("transaction_type__code") or row.get("transaction_type__transaction_category") or ""
        elif group_by == "payment_method":
            group_id = str(row.get("payment_method_id") or "")
            label = row.get("payment_method__name") or "Unknown method"
            sub_label = row.get("payment_method__code") or ""
        else:
            month_value = row.get("month")
            group_id = month_value.strftime("%Y-%m") if month_value else ""
            label = month_value.strftime("%b %Y") if month_value else "Unknown month"
            sub_label = group_id

        income = float(row.get("income") or 0)
        expense = float(row.get("expense") or 0)
        income_count = row.get("income_count") or 0
        expense_count = row.get("expense_count") or 0

        return {
            "group_id": group_id,
            "group_label": label,
            "group_sub_label": sub_label,
            "bank_account_number": account_number if group_by == "bank_account" else "",
            "bank_name": bank_name if group_by == "bank_account" else "",
            "income": income,
            "expense": expense,
            "net": income - expense,
            "balance": income - expense,
            "income_count": income_count,
            "expense_count": expense_count,
            "transaction_count": income_count + expense_count,
        }

    @staticmethod
    def _serialize_transaction(txn, currency_symbol: str) -> dict:
        category = getattr(txn.transaction_type, "transaction_category", "") or ""
        return {
            "id": str(txn.id),
            "reference_number": txn.reference_number or "",
            "transaction_date": txn.transaction_date.isoformat() if txn.transaction_date else "",
            "category": category,
            "transaction_type_code": getattr(txn.transaction_type, "code", "") or "",
            "transaction_type_name": getattr(txn.transaction_type, "name", "") or "",
            "description": txn.description or "",
            "bank_account_name": getattr(txn.bank_account, "account_name", "") or "",
            "bank_account_number": getattr(txn.bank_account, "account_number", "") or "",
            "bank_name": getattr(txn.bank_account, "bank_name", "") or "",
            "payment_method_name": getattr(txn.payment_method, "name", "") or "",
            "payer_payee": txn.payer_payee or "",
            "source_reference": txn.source_reference or "",
            "amount": str(txn.amount or 0),
            "base_amount": str(txn.base_amount or 0),
            "currency_symbol": txn.currency.symbol if txn.currency else currency_symbol,
            "base_currency_symbol": currency_symbol,
            "status": txn.status,
        }

    def get(self, request):
        group_by = self._parse_group_by(request.query_params.get("group_by"))

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        status_param = (request.query_params.get("status") or "approved").strip().lower()
        fmt = request.query_params.get("export") or request.query_params.get("format")

        filter_params: dict[str, str] = {}
        if start_date:
            filter_params["start_date"] = start_date
        if end_date:
            filter_params["end_date"] = end_date
        if status_param and status_param != "all":
            filter_params["status"] = status_param
        bank_account = request.query_params.get("bank_account")
        if bank_account:
            filter_params["bank_account"] = bank_account

        category_filter = (request.query_params.get("category") or "").strip().lower()
        if category_filter in {"income", "expense"}:
            filter_params["category"] = category_filter

        base_currency = serialize_currency(get_tenant_base_currency())
        currency_symbol = base_currency["symbol"] or base_currency["code"] or "$"

        cash_scope_queryset = apply_cash_transaction_list_filters(
            approved_cash_queryset(),
            filter_params,
        )
        net_cash_movement = float(
            sum_approved_cash_net(
                filter_cash_by_period(cash_scope_queryset, start_date, end_date)
            )
        )

        queryset = apply_cash_transaction_list_filters(
            AccountingCashTransaction.objects.select_related(
                "transaction_type",
                "bank_account",
                "payment_method",
                "currency",
            ),
            filter_params,
        ).filter(LEDGER_CATEGORIES).order_by("-transaction_date", "-created_at")

        if group_by is None:
            groups_by = {
                key: self._compute_groups(queryset, key) for key in GROUP_BY_OPTIONS
            }
            groups = groups_by["bank_account"]
            response_group_by = "all"
        else:
            groups_by = None
            groups = self._compute_groups(queryset, group_by)
            response_group_by = group_by

        totals = queryset.aggregate(
            income_total=Sum(Abs(F("base_amount")), filter=INCOME),
            expense_total=Sum(Abs(F("base_amount")), filter=EXPENSE),
            income_count=Count("id", filter=INCOME),
            expense_count=Count("id", filter=EXPENSE),
            transaction_count=Count("id"),
        )
        income_total = float(totals["income_total"] or 0)
        expense_total = float(totals["expense_total"] or 0)
        operating_net = float(sum_period_income_minus_expense(queryset))

        results = []
        for txn in queryset:
            results.append(self._serialize_transaction(txn, currency_symbol))

        payload = {
            "start_date": start_date or "",
            "end_date": end_date or "",
            "status": status_param,
            "group_by": response_group_by,
            "count": len(results),
            "summary": {
                "income_total": income_total,
                "expense_total": expense_total,
                "net_total": operating_net,
                "operating_net": operating_net,
                "net_cash_movement": net_cash_movement,
                "balance_total": operating_net,
                "income_count": totals["income_count"] or 0,
                "expense_count": totals["expense_count"] or 0,
                "transaction_count": totals["transaction_count"] or 0,
                "currency_symbol": currency_symbol,
                "base_currency_code": base_currency["code"],
            },
            "groups": groups,
            "results": results,
        }
        if groups_by is not None:
            payload["groups_by"] = groups_by

        if fmt == "xlsx":
            return self._export_xlsx(payload)
        if fmt == "pdf":
            return self._export_pdf(request, payload)

        return Response(payload)

    def _export_xlsx(self, payload: dict) -> HttpResponse:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        from ..utils.export_helpers import apply_xlsx_cell_style

        currency = payload.get("summary", {}).get("currency_symbol") or "$"

        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = "Overview"
        ws_overview["A1"] = "Income & Expense Summary"
        ws_overview["A1"].font = Font(bold=True, size=14)
        ws_overview["A2"] = f"Period: {payload['start_date']} to {payload['end_date']}"
        ws_overview["A3"] = f"Status: {payload['status']}"
        overview_rows = [
            ("Total Income", payload["summary"]["income_total"]),
            ("Total Expense", payload["summary"]["expense_total"]),
            ("Operating Net", payload["summary"]["operating_net"]),
            ("Net Cash Movement", payload["summary"]["net_cash_movement"]),
        ]
        for row_offset, (label, value) in enumerate(overview_rows, 5):
            ws_overview[f"A{row_offset}"] = label
            cell = ws_overview[f"B{row_offset}"]
            cell.value = value
            apply_xlsx_cell_style(cell, label, value, currency)

        grouped_sections = payload.get("groups_by") or {
            payload.get("group_by", "bank_account"): payload.get("groups", [])
        }
        section_titles = {
            "bank_account": "By Bank Account",
            "transaction_type": "By Transaction Type",
            "payment_method": "By Payment Method",
            "month": "By Month",
        }

        for group_key, rows in grouped_sections.items():
            title = section_titles.get(group_key, group_key.replace("_", " ").title())
            ws = wb.create_sheet(title[:31])
            headers = ["Group", "Detail", "Account No.", "Income", "Expense", "Balance", "Transactions"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            for row_offset, row in enumerate(rows, 2):
                ws.cell(row=row_offset, column=1, value=row["group_label"])
                ws.cell(row=row_offset, column=2, value=row["group_sub_label"])
                ws.cell(row=row_offset, column=3, value=row.get("bank_account_number") or "")
                for col_idx, header in enumerate(headers[3:6], 4):
                    value = (
                        row["income"]
                        if col_idx == 4
                        else row["expense"]
                        if col_idx == 5
                        else row.get("balance", row["net"])
                    )
                    cell = ws.cell(row=row_offset, column=col_idx, value=value)
                    apply_xlsx_cell_style(cell, header, value, currency)
                ws.cell(row=row_offset, column=7, value=row["transaction_count"])

        ws_detail = wb.create_sheet("Transactions")
        detail_headers = [
            "Reference",
            "Date",
            "Category",
            "Type",
            "Description",
            "Bank Account",
            "Account No.",
            "Payment Method",
            "Amount",
            "Status",
        ]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_offset, row in enumerate(payload["results"], 2):
            values = [
                row["reference_number"],
                row["transaction_date"],
                row["category"],
                row["transaction_type_name"] or row["transaction_type_code"],
                row["description"],
                row["bank_account_name"],
                row.get("bank_account_number") or "",
                row["payment_method_name"],
                float(row["amount"]),
                row["status"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_detail.cell(row=row_offset, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = Font(size=9)
                if col_idx == 9:
                    apply_xlsx_cell_style(cell, "Amount", value, currency)

        col_widths = [16, 12, 10, 18, 36, 20, 14, 16, 12, 12]
        for col_idx, width in enumerate(col_widths, 1):
            ws_detail.column_dimensions[get_column_letter(col_idx)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"income-expense-summary-{payload['start_date']}-to-{payload['end_date']}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_pdf(self, request, payload: dict) -> HttpResponse:
        from ..utils.export_helpers import build_pdf_response, format_currency_display

        currency = payload.get("summary", {}).get("currency_symbol") or "$"
        income_total = format_currency_display(payload["summary"]["income_total"], currency)
        expense_total = format_currency_display(payload["summary"]["expense_total"], currency)

        headers = ["Reference", "Date", "Category", "Type", "Description", "Amount", "Status"]
        rows = [
            [
                row["reference_number"],
                row["transaction_date"],
                row["category"],
                row.get("transaction_type_name") or row.get("transaction_type_code") or "",
                row["description"],
                float(row["amount"]),
                row["status"],
            ]
            for row in payload["results"]
        ]
        return build_pdf_response(
            request=request,
            filename=f"income-expense-summary-{payload['start_date']}-to-{payload['end_date']}.pdf",
            title="Income & Expense Summary",
            subtitle=(
                f"Period: {payload['start_date']} to {payload['end_date']} | "
                f"Income: {income_total} | Expense: {expense_total}"
            ),
            headers=headers,
            rows=rows,
            currency_symbol=currency,
        )
