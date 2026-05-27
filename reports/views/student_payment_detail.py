"""
Student payment detail report — one row per cash transaction payment.
"""

import io
from datetime import date
from decimal import Decimal, InvalidOperation

from django.db.models import Prefetch, Q
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from ..access_policies import ReportsAccessPolicy
from .finance import FinanceReportView


class StudentPaymentDetailReportView(APIView):
    """Individual student payment lines with transaction reference numbers."""

    permission_classes = [ReportsAccessPolicy]

    @staticmethod
    def _parse_decimal_param(value: str | None) -> Decimal | None:
        return FinanceReportView._parse_decimal_param(value)

    @staticmethod
    def _read_multi_query_values(request, key: str) -> list[str]:
        return FinanceReportView._read_multi_query_values(request, key)

    @staticmethod
    def _resolve_student_for_transaction(transaction, enrollment_map):
        student = transaction.student
        if student:
            return student

        for allocation in transaction.bill_allocations.all():
            bill = allocation.student_bill
            if bill and bill.student_id:
                return bill.student
        return None

    def get(self, request):
        from academics.models import AcademicYear
        from accounting.models import (
            AccountingCashTransaction,
            AccountingStudentPaymentAllocation,
        )
        from accounting.services.post_all import build_student_payment_list_filter
        from students.models import Enrollment

        academic_year_id = request.query_params.get("academic_year_id")
        grade_level_ids = self._read_multi_query_values(request, "grade_level_id")
        section_ids = self._read_multi_query_values(request, "section_id")
        student_query = (request.query_params.get("student") or "").strip()
        reference_query = (request.query_params.get("reference") or "").strip()
        status_param = (request.query_params.get("status") or "").strip()
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        amount_min = self._parse_decimal_param(request.query_params.get("amount_min"))
        amount_max = self._parse_decimal_param(request.query_params.get("amount_max"))
        fmt = request.query_params.get("export") or request.query_params.get("format")

        if academic_year_id:
            try:
                academic_year = AcademicYear.objects.get(id=academic_year_id)
            except AcademicYear.DoesNotExist:
                return Response({"detail": "Academic year not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            academic_year = AcademicYear.objects.filter(current=True).first()
            if not academic_year:
                return Response({"detail": "No current academic year found."}, status=status.HTTP_400_BAD_REQUEST)

        queryset = (
            AccountingCashTransaction.objects.filter(
                build_student_payment_list_filter(),
                transaction_date__gte=academic_year.start_date,
                transaction_date__lte=academic_year.end_date,
            )
            .select_related(
                "student",
                "payment_method",
                "bank_account",
                "currency",
                "transaction_type",
            )
            .prefetch_related(
                Prefetch(
                    "bill_allocations",
                    queryset=AccountingStudentPaymentAllocation.objects.select_related(
                        "student_bill__student"
                    ),
                )
            )
            .order_by("-transaction_date", "-created_at")
        )

        if start_date:
            queryset = queryset.filter(transaction_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(transaction_date__lte=end_date)
        if reference_query:
            queryset = queryset.filter(reference_number__icontains=reference_query)
        if status_param:
            queryset = queryset.filter(status=status_param)
        if amount_min is not None:
            queryset = queryset.filter(amount__gte=amount_min)
        if amount_max is not None:
            queryset = queryset.filter(amount__lte=amount_max)
        if student_query:
            queryset = queryset.filter(
                Q(student__id_number__icontains=student_query)
                | Q(student__first_name__icontains=student_query)
                | Q(student__last_name__icontains=student_query)
                | Q(payer_payee__icontains=student_query)
                | Q(source_reference__icontains=student_query)
            )

        if grade_level_ids or section_ids:
            enrollment_filter = Q(academic_year=academic_year)
            if grade_level_ids:
                enrollment_filter &= Q(grade_level_id__in=grade_level_ids)
            if section_ids:
                enrollment_filter &= Q(section_id__in=section_ids)
            scoped_student_ids = Enrollment.objects.filter(enrollment_filter).values_list(
                "student_id", flat=True
            )
            queryset = queryset.filter(
                Q(student_id__in=scoped_student_ids)
                | Q(bill_allocations__student_bill__student_id__in=scoped_student_ids)
            ).distinct()

        transactions = list(queryset)

        student_ids = set()
        for txn in transactions:
            student = self._resolve_student_for_transaction(txn, {})
            if student:
                student_ids.add(student.id)

        enrollments = Enrollment.objects.filter(
            academic_year=academic_year,
            student_id__in=student_ids,
        ).select_related("grade_level", "section")
        enrollment_map = {enrollment.student_id: enrollment for enrollment in enrollments}

        from accounting.models import AccountingStudentBill

        student_bills = AccountingStudentBill.objects.filter(
            academic_year=academic_year,
            student_id__in=student_ids,
        )
        bill_map = {bill.student_id: bill for bill in student_bills}
        student_balance_map = FinanceReportView._build_student_balance_map(
            student_ids,
            academic_year,
        )

        results = []
        for txn in transactions:
            student = self._resolve_student_for_transaction(txn, enrollment_map)
            enrollment = enrollment_map.get(student.id) if student else None
            bill = bill_map.get(student.id) if student else None
            balance_info = student_balance_map.get(str(student.id), {}) if student else {}

            results.append(
                {
                    "id": str(txn.id),
                    "reference_number": txn.reference_number or "",
                    "transaction_date": str(txn.transaction_date) if txn.transaction_date else "",
                    "student_id": student.id_number if student else "",
                    "student_name": student.get_full_name() if student else (txn.payer_payee or ""),
                    "grade_level_id": (
                        str(enrollment.grade_level_id)
                        if enrollment and enrollment.grade_level_id
                        else ""
                    ),
                    "grade_level": (
                        enrollment.grade_level.name
                        if enrollment and enrollment.grade_level
                        else ""
                    ),
                    "section_id": (
                        str(enrollment.section_id)
                        if enrollment and enrollment.section_id
                        else ""
                    ),
                    "section": (
                        enrollment.section.name if enrollment and enrollment.section else ""
                    ),
                    "gross": float(bill.gross_amount) if bill else 0,
                    "concession": float(bill.concession_amount) if bill else 0,
                    "net": float(bill.net_amount) if bill else 0,
                    "balance": balance_info.get(
                        "balance_total",
                        float(bill.outstanding_amount) if bill else 0,
                    ),
                    "amount": float(txn.amount or 0),
                    "currency": txn.currency.symbol if txn.currency else "$",
                    "payment_method": txn.payment_method.name if txn.payment_method else "",
                    "bank_account": txn.bank_account.account_name if txn.bank_account else "",
                    "transaction_type": txn.transaction_type.name if txn.transaction_type else "",
                    "status": txn.status or "",
                    "description": txn.description or "",
                    "payer_payee": txn.payer_payee or "",
                }
            )

        totals = {
            "payment_count": len(results),
            "total_amount": sum(row["amount"] for row in results),
        }

        if fmt == "xlsx":
            return self._export_xlsx(results, totals, academic_year)

        return Response(
            {
                "count": len(results),
                "academic_year": {"id": str(academic_year.id), "name": academic_year.name},
                "results": results,
                "totals": totals,
            }
        )

    def _export_xlsx(self, results, totals, academic_year):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        from ..utils.export_helpers import apply_xlsx_cell_style, resolve_export_currency

        currency = resolve_export_currency()

        today = date.today()
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Payments"

        ws.merge_cells("A1:P1")
        ws["A1"] = "Student Payment Detail Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:P2")
        ws["A2"] = f"Academic Year {academic_year.name}"
        ws["A2"].font = Font(bold=True, size=11)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:P3")
        ws["A3"] = f"Report Date: {today.strftime('%A, %B %d, %Y')}"
        ws["A3"].alignment = Alignment(horizontal="center")

        header_row = 5
        headers = [
            "Reference",
            "Date",
            "Student ID",
            "Student Name",
            "Grade Level",
            "Section",
            "Gross",
            "Concession",
            "Net",
            "Paid",
            "Balance",
            "Payment Method",
            "Bank Account",
            "Type",
            "Status",
            "Description",
        ]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        currency_col = 11
        currency_columns = {7, 8, 9, 10, 11}
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        data_start = header_row + 1
        for row_offset, row in enumerate(results):
            excel_row = data_start + row_offset
            values = [
                row["reference_number"],
                row["transaction_date"],
                row["student_id"],
                row["student_name"],
                row["grade_level"],
                row["section"],
                row["gross"],
                row["concession"],
                row["net"],
                row["amount"],
                row["balance"],
                row["payment_method"],
                row["bank_account"],
                row["transaction_type"],
                row["status"],
                row["description"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = Font(size=9)
                if col_idx in currency_columns:
                    apply_xlsx_cell_style(cell, "Amount", value, currency)

        total_row = data_start + len(results)
        label_cell = ws.cell(
            row=total_row,
            column=1,
            value=f"Totals for {totals['payment_count']} payments",
        )
        label_cell.font = Font(bold=True, size=9)
        ws.merge_cells(f"A{total_row}:F{total_row}")

        total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        total_font = Font(bold=True, size=9)
        for col_idx in range(1, 17):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = thin_border

        amount_cell = ws.cell(row=total_row, column=currency_col, value=totals["total_amount"])
        apply_xlsx_cell_style(amount_cell, "Amount", totals["total_amount"], currency)

        col_widths = [16, 12, 12, 28, 14, 14, 12, 12, 12, 12, 12, 16, 18, 16, 12, 28]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = f"A{data_start}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_year = academic_year.name.replace(" ", "-").replace("/", "-").lower()
        filename = f"student-payments-{safe_year}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
