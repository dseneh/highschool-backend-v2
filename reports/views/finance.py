"""
Finance Reports Views

Fees Payment SITREP report sourced from the accounting module.
"""

import io
from datetime import date
from decimal import Decimal, InvalidOperation

from django.db.models import Prefetch, Q, Sum
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from ..access_policies import ReportsAccessPolicy

PAYMENT_STATUS_LABELS = {
    "fully_paid": "Fully Paid",
    "payment_current": "Payment Current",
    "delinquent": "Delinquent",
    "overpaid": "Overpaid",
    "credit": "Overpaid",
}


class FinanceReportView(APIView):
    """Fees Payment SITREP - sourced from the accounting module."""

    permission_classes = [ReportsAccessPolicy]

    @staticmethod
    def _parse_decimal_param(value: str | None) -> Decimal | None:
        if value is None:
            return None
        normalized = str(value).strip()
        if not normalized:
            return None
        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _normalize_payment_status_filters(raw_values: list[str]) -> set[str]:
        labels: set[str] = set()
        for raw in raw_values:
            key = raw.strip().lower().replace(" ", "_")
            if not key or key == "all":
                continue
            label = PAYMENT_STATUS_LABELS.get(key)
            if label:
                labels.add(label)
            elif raw.strip():
                labels.add(raw.strip())
        return labels

    @staticmethod
    def _derive_payment_status_label(
        balance: float,
        *,
        has_overdue_bill: bool,
        amt_due_todate: float,
        total_paid: float,
    ) -> str:
        if balance < 0:
            return "Overpaid"
        if balance == 0:
            return "Fully Paid"
        if has_overdue_bill:
            return "Delinquent"
        if amt_due_todate > 0 and total_paid < amt_due_todate:
            return "Delinquent"
        return "Payment Current"

    @staticmethod
    def _passes_amount_filters(
        row: dict,
        *,
        balance_min: Decimal | None,
        balance_max: Decimal | None,
        total_paid_min: Decimal | None,
        total_paid_max: Decimal | None,
        net_bill_min: Decimal | None,
        net_bill_max: Decimal | None,
    ) -> bool:
        checks = [
            (balance_min, row["balance"], lambda value, bound: value >= bound),
            (balance_max, row["balance"], lambda value, bound: value <= bound),
            (total_paid_min, row["total_paid"], lambda value, bound: value >= bound),
            (total_paid_max, row["total_paid"], lambda value, bound: value <= bound),
            (net_bill_min, row["net_bill"], lambda value, bound: value >= bound),
            (net_bill_max, row["net_bill"], lambda value, bound: value <= bound),
        ]
        for bound, value, comparator in checks:
            if bound is None:
                continue
            if not comparator(float(value), float(bound)):
                return False
        return True

    @staticmethod
    def _build_student_balance_map(student_ids, academic_year) -> dict[str, dict[str, float]]:
        if not student_ids:
            return {}

        from students.models import Student
        from students.services.balance import annotate_student_balance_totals

        balance_rows = annotate_student_balance_totals(
            Student.objects.filter(id__in=student_ids),
            academic_year=academic_year,
        )
        return {
            str(student.id): {
                "paid_total": float(student.paid_total or 0),
                "billed_total": float(student.billed_total or 0),
                "balance_total": float(student.balance_total or 0),
            }
            for student in balance_rows
        }

    @staticmethod
    def _read_multi_query_values(request, key: str) -> list[str]:
        values: list[str] = []

        for raw in request.query_params.getlist(key):
            if raw is None:
                continue
            for part in str(raw).split(","):
                value = part.strip()
                if value:
                    values.append(value)

        if not values:
            single = request.query_params.get(key)
            if single:
                for part in str(single).split(","):
                    value = part.strip()
                    if value:
                        values.append(value)

        # Preserve order while removing duplicates
        return list(dict.fromkeys(values))

    @staticmethod
    def _build_student_paid_map(bills_list, academic_year) -> dict[str, float]:
        """Approved paid amount per student for the selected academic year.

        Includes transactions matched directly by student, source_reference fallback,
        and bill allocations. Allocation-linked rows are counted for the selected
        academic year even when transaction_date falls outside year bounds.
        """
        from accounting.models import AccountingCashTransaction
        from accounting.services.post_all import build_student_payment_list_filter

        if not bills_list:
            return {}

        student_ids = {bill.student_id for bill in bills_list if bill.student_id}
        if not student_ids:
            return {}

        # Resolve source_reference fallbacks to student IDs.
        reference_to_student: dict[str, str] = {}
        for bill in bills_list:
            student = bill.student
            if not student:
                continue
            student_key = str(student.id)
            reference_to_student[str(student.id)] = student_key
            if getattr(student, "id_number", None):
                reference_to_student[str(student.id_number)] = student_key
            if getattr(student, "prev_id_number", None):
                reference_to_student[str(student.prev_id_number)] = student_key

        transactions = (
            AccountingCashTransaction.objects.filter(
                build_student_payment_list_filter(),
                status=AccountingCashTransaction.TransactionStatus.APPROVED,
            )
            .filter(
                Q(
                    transaction_date__gte=academic_year.start_date,
                    transaction_date__lte=academic_year.end_date,
                )
                | Q(bill_allocations__student_bill__academic_year=academic_year)
            )
            .filter(
                Q(student_id__in=student_ids)
                | Q(source_reference__in=list(reference_to_student.keys()))
                | Q(bill_allocations__student_bill__student_id__in=student_ids)
            )
            .select_related("student")
            .prefetch_related("bill_allocations__student_bill")
            .distinct()
        )

        paid_map: dict[str, float] = {str(student_id): 0.0 for student_id in student_ids}
        for tx in transactions:
            student_key = None

            if tx.student_id and tx.student_id in student_ids:
                student_key = str(tx.student_id)
            elif tx.source_reference:
                student_key = reference_to_student.get(str(tx.source_reference))

            if not student_key:
                for allocation in tx.bill_allocations.all():
                    bill = allocation.student_bill
                    if (
                        bill
                        and bill.academic_year_id == academic_year.id
                        and bill.student_id in student_ids
                    ):
                        student_key = str(bill.student_id)
                        break

            if not student_key:
                continue

            # Always aggregate in base currency so mixed-currency payments
            # are comparable with billing totals.
            paid_value = tx.base_amount if tx.base_amount is not None else tx.amount
            paid_map[student_key] = paid_map.get(student_key, 0.0) + float(paid_value or 0)

        return paid_map

    @staticmethod
    def _compute_student_payment_total(academic_year) -> float:
        """Approved student-payment total in base currency for the academic year."""
        from accounting.models import AccountingCashTransaction
        from accounting.services.post_all import build_student_payment_list_filter

        transactions = (
            AccountingCashTransaction.objects.filter(
                build_student_payment_list_filter(),
                status=AccountingCashTransaction.TransactionStatus.APPROVED,
            )
            .filter(
                Q(
                    transaction_date__gte=academic_year.start_date,
                    transaction_date__lte=academic_year.end_date,
                )
                | Q(bill_allocations__student_bill__academic_year=academic_year)
            )
            .distinct()
            .only("base_amount", "amount")
        )

        total = 0.0
        for tx in transactions:
            paid_value = tx.base_amount if tx.base_amount is not None else tx.amount
            total += float(paid_value or 0)

        return total

    def get(self, request):
        from academics.models import AcademicYear
        from accounting.models import (
            AccountingConcession,
            AccountingInstallmentLine,
            AccountingInstallmentPlan,
            AccountingStudentBill,
            AccountingStudentBillLine,
        )

        academic_year_id = request.query_params.get("academic_year_id")
        grade_level_ids = self._read_multi_query_values(request, "grade_level_id")
        section_ids = self._read_multi_query_values(request, "section_id")
        payment_statuses = self._normalize_payment_status_filters(
            self._read_multi_query_values(request, "payment_status")
        )
        student_query = (request.query_params.get("student") or "").strip()
        balance_min = self._parse_decimal_param(request.query_params.get("balance_min"))
        balance_max = self._parse_decimal_param(request.query_params.get("balance_max"))
        total_paid_min = self._parse_decimal_param(request.query_params.get("total_paid_min"))
        total_paid_max = self._parse_decimal_param(request.query_params.get("total_paid_max"))
        net_bill_min = self._parse_decimal_param(request.query_params.get("net_bill_min"))
        net_bill_max = self._parse_decimal_param(request.query_params.get("net_bill_max"))
        # NOTE: DRF reserves `format` for renderer negotiation.
        # Use `export=xlsx` to avoid 404/negotiation issues, but keep `format` as fallback.
        fmt = request.query_params.get("export") or request.query_params.get("format")

        # Resolve academic year
        if academic_year_id:
            try:
                academic_year = AcademicYear.objects.get(id=academic_year_id)
            except AcademicYear.DoesNotExist:
                return Response({"detail": "Academic year not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            academic_year = AcademicYear.objects.filter(current=True).first()
            if not academic_year:
                return Response({"detail": "No current academic year found."}, status=status.HTTP_400_BAD_REQUEST)

        # Get active installment plan for this academic year
        today = date.today()
        installment_plan = (
            AccountingInstallmentPlan.objects.filter(academic_year=academic_year, is_active=True)
            .prefetch_related(
                Prefetch("lines", queryset=AccountingInstallmentLine.objects.order_by("sequence"))
            )
            .first()
        )

        installment_lines = list(installment_plan.lines.all()) if installment_plan else []

        # Cumulative expected percentage up to today
        cumulative_pct = (
            sum(float(line.percentage) for line in installment_lines if line.due_date <= today) / 100.0
        )

        # Name of the most recent installment that has come due
        current_installment_name = ""
        for line in sorted(installment_lines, key=lambda l: l.due_date):
            if line.due_date <= today:
                current_installment_name = line.name

        # Query student bills
        bills = (
            AccountingStudentBill.objects.filter(academic_year=academic_year)
            .exclude(status=AccountingStudentBill.BillStatus.CANCELLED)
            .select_related(
                "student",
                "enrollment",
                "enrollment__section",
                "grade_level",
                "currency",
            )
            .prefetch_related(
                Prefetch(
                    "lines",
                    queryset=AccountingStudentBillLine.objects.select_related("fee_item"),
                )
            )
            .order_by(
                "grade_level__level",
                "enrollment__section__name",
                "student__last_name",
                "student__first_name",
            )
        )

        if grade_level_ids:
            bills = bills.filter(grade_level_id__in=grade_level_ids)
        if section_ids:
            bills = bills.filter(enrollment__section_id__in=section_ids)
        if student_query:
            bills = bills.filter(
                Q(student__id_number__icontains=student_query)
                | Q(student__first_name__icontains=student_query)
                | Q(student__last_name__icontains=student_query)
            )

        bills_list = list(bills)
        scoped_student_ids = {bill.student_id for bill in bills_list if bill.student_id}

        live_concession_qs = AccountingConcession.objects.filter(
            academic_year=academic_year,
            is_active=True,
        )
        if scoped_student_ids:
            live_concession_qs = live_concession_qs.filter(student_id__in=scoped_student_ids)

        live_concession_map = {
            str(row["student_id"]): float(row["total"] or 0)
            for row in live_concession_qs.values("student_id").annotate(total=Sum("computed_amount"))
        }
        use_live_concessions = bool(live_concession_map)

        student_balance_map = self._build_student_balance_map(
            scoped_student_ids,
            academic_year,
        )
        student_paid_map = self._build_student_paid_map(bills_list, academic_year)

        student_rows = {}
        for bill in bills_list:
            gross_amount = float(bill.gross_amount or 0)
            net_amount = float(bill.net_amount or 0)
            derived_concession = round(gross_amount - net_amount, 2)
            tuition = sum(
                float(line.line_amount)
                for line in bill.lines.all()
                if line.fee_item.category == "tuition"
            )
            adm_fees = sum(
                float(line.line_amount)
                for line in bill.lines.all()
                if line.fee_item.category != "tuition"
            )

            student_key = str(bill.student_id)
            if student_key not in student_rows:
                enrolled_as = ""
                enrolled_as_display = ""
                if bill.enrollment:
                    enrolled_as = bill.enrollment.enrolled_as or ""
                    enrolled_as_display = (
                        bill.enrollment.get_enrolled_as_display()
                        if hasattr(bill.enrollment, "get_enrolled_as_display")
                        else enrolled_as.capitalize()
                    )

                student_rows[student_key] = {
                    "id": student_key,
                    "student_id": bill.student.id_number,
                    "student_name": bill.student.get_full_name(),
                    "grade_level_id": str(bill.grade_level_id) if bill.grade_level_id else "",
                    "grade_level": bill.grade_level.name if bill.grade_level else "",
                    "section_id": (
                        str(bill.enrollment.section_id)
                        if bill.enrollment and bill.enrollment.section_id
                        else ""
                    ),
                    "section": (
                        bill.enrollment.section.name
                        if bill.enrollment and bill.enrollment.section
                        else ""
                    ),
                    "enrolled_as": enrolled_as,
                    "enrolled_as_display": enrolled_as_display,
                    "tuition": 0.0,
                    "adm_fees": 0.0,
                    "total_bill": 0.0,
                    "concession": 0.0,
                    "net_bill": 0.0,
                    "current_installment": current_installment_name,
                    "currency": bill.currency.symbol if bill.currency else "$",
                    "_paid_fallback": 0.0,
                    "_concession_fallback": 0.0,
                    "_has_overdue_bill": False,
                }

            row = student_rows[student_key]
            row["tuition"] += tuition
            row["adm_fees"] += adm_fees
            row["total_bill"] += gross_amount
            row["net_bill"] += net_amount
            row["_paid_fallback"] += float(bill.paid_amount or 0)
            row["_concession_fallback"] += float(bill.concession_amount or 0)
            row["_has_overdue_bill"] = row["_has_overdue_bill"] or (
                bill.status == bill.BillStatus.OVERDUE
            )

        results = []
        for student_key, row in student_rows.items():
            total_bill = float(row["total_bill"])
            if use_live_concessions:
                concession = min(total_bill, max(0.0, float(live_concession_map.get(student_key, 0.0))))
                net_bill = round(max(0.0, total_bill - concession), 2)
            else:
                concession = float(row.get("_concession_fallback", 0.0))
                net_bill = float(row["net_bill"])

            balance_info = student_balance_map.get(student_key, {})
            paid = student_paid_map.get(
                student_key,
                balance_info.get("paid_total", float(row["_paid_fallback"])),
            )
            balance = round(net_bill - paid, 2)
            credit_amount = round(abs(min(0.0, balance)), 2)

            amt_due_todate = round(net_bill * cumulative_pct, 2) if cumulative_pct > 0 else 0
            pct_paid_due = round((paid / amt_due_todate * 100), 1) if amt_due_todate > 0 else 0
            pct_paid_net = round((paid / net_bill * 100), 1) if net_bill > 0 else 0

            status_label = self._derive_payment_status_label(
                balance,
                has_overdue_bill=bool(row.get("_has_overdue_bill")),
                amt_due_todate=amt_due_todate,
                total_paid=paid,
            )

            results.append(
                {
                    "id": row["id"],
                    "student_id": row["student_id"],
                    "student_name": row["student_name"],
                    "grade_level_id": row["grade_level_id"],
                    "grade_level": row["grade_level"],
                    "section_id": row["section_id"],
                    "section": row["section"],
                    "enrolled_as": row["enrolled_as"],
                    "enrolled_as_display": row["enrolled_as_display"],
                    "tuition": row["tuition"],
                    "adm_fees": row["adm_fees"],
                    "total_bill": total_bill,
                    "concession": concession,
                    "net_bill": net_bill,
                    "current_installment": row["current_installment"],
                    "amt_due_todate": amt_due_todate,
                    "total_paid": paid,
                    "balance": balance,
                    "credit_amount": credit_amount,
                    "pct_paid_due": pct_paid_due,
                    "pct_paid_net": pct_paid_net,
                    "status": status_label,
                    "currency": row["currency"],
                }
            )

        if payment_statuses:
            results = [row for row in results if row["status"] in payment_statuses]

        results = [
            row
            for row in results
            if self._passes_amount_filters(
                row,
                balance_min=balance_min,
                balance_max=balance_max,
                total_paid_min=total_paid_min,
                total_paid_max=total_paid_max,
                net_bill_min=net_bill_min,
                net_bill_max=net_bill_max,
            )
        ]

        status_counts = {
            "Fully Paid": 0,
            "Payment Current": 0,
            "Delinquent": 0,
            "Overpaid": 0,
        }
        for row in results:
            status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1

        totals = {
            "student_count": len(results),
            "tuition": sum(r["tuition"] for r in results),
            "adm_fees": sum(r["adm_fees"] for r in results),
            "total_bill": sum(r["total_bill"] for r in results),
            "concession": sum(r["concession"] for r in results),
            "net_bill": sum(r["net_bill"] for r in results),
            "amt_due_todate": sum(r["amt_due_todate"] for r in results),
            "total_paid": sum(r["total_paid"] for r in results),
            "balance": 0.0,
            "outstanding_balance": 0.0,
            "credit_total": sum(r["credit_amount"] for r in results),
            "overpaid_count": status_counts.get("Overpaid", 0),
            "status_counts": status_counts,
        }
        if not grade_level_ids and not section_ids and not student_query:
            # Keep top-level paid total aligned with the student-payment
            # transaction source used by the cash-transactions module.
            totals["total_paid"] = self._compute_student_payment_total(academic_year)
        totals["balance"] = round(totals["net_bill"] - totals["total_paid"], 2)
        totals["outstanding_balance"] = max(0.0, totals["balance"])
        total_net = totals["net_bill"]
        total_paid_sum = totals["total_paid"]
        totals["pct_paid_net"] = round((total_paid_sum / total_net * 100), 1) if total_net > 0 else 0

        if fmt == "xlsx":
            return self._export_xlsx(results, totals, academic_year)
        if fmt == "pdf":
            return self._export_pdf(request, results, totals, academic_year)

        return Response(
            {
                "count": len(results),
                "academic_year": {"id": str(academic_year.id), "name": academic_year.name},
                "current_installment": current_installment_name,
                "results": results,
                "totals": totals,
            }
        )

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def _export_xlsx(self, results, totals, academic_year):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        money_fmt = "#,##0.00"

        today = date.today()

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Billing Summary"

        # Header rows
        ws.merge_cells("A1:Q1")
        ws["A1"] = "Situation Report, Student Fees Payment"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:Q2")
        ws["A2"] = f"Academic Year {academic_year.name}"
        ws["A2"].font = Font(bold=True, size=11)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:Q3")
        ws["A3"] = f"Report Date: {today.strftime('%A, %B %d, %Y')}"
        ws["A3"].alignment = Alignment(horizontal="center")

        # Column headers (row 5)
        HEADER_ROW = 5
        headers = [
            "Student ID",
            "Student Name",
            "Grade Level",
            "Section",
            "En. As",
            "Tuition",
            "Adm Fees",
            "Total Bill",
            "Concession",
            "Net Bill",
            "Current Instalmt",
            "Amt Due Todate",
            "Total Paid",
            "Balance",
            "% Paid, Due Tdte",
            "% Paid, Tot Bill",
            "Status",
        ]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Column groups
        currency_cols = {6, 7, 8, 9, 10, 12, 13, 14}
        pct_cols = {15, 16}
        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        status_colors = {
            "Fully Paid": "C6EFCE",
            "Payment Current": "DDEBF7",
            "Delinquent": "FFCCCC",
            "Overpaid": "D9E1F2",
        }

        # Data rows
        DATA_START = HEADER_ROW + 1
        for row_offset, row in enumerate(results):
            excel_row = DATA_START + row_offset
            values = [
                row["student_id"],
                row["student_name"],
                row["grade_level"],
                row["section"],
                row.get("enrolled_as_display") or row.get("enrolled_as", ""),
                row["tuition"],
                row["adm_fees"],
                row["total_bill"],
                row["concession"],
                row["net_bill"],
                row["current_installment"],
                row["amt_due_todate"],
                row["total_paid"],
                row["balance"],
                (row["pct_paid_due"] / 100) if row["pct_paid_due"] else 0,
                (row["pct_paid_net"] / 100) if row["pct_paid_net"] else 0,
                row["status"],
            ]
            status_fill_color = status_colors.get(row["status"])

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = Font(size=9)
                if col_idx in currency_cols:
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in pct_cols:
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                if status_fill_color and col_idx == 17:
                    cell.fill = PatternFill(
                        start_color=status_fill_color,
                        end_color=status_fill_color,
                        fill_type="solid",
                    )

        # Totals row
        total_row = DATA_START + len(results)
        label_cell = ws.cell(
            row=total_row,
            column=1,
            value=f"Totals for {totals['student_count']} students",
        )
        label_cell.font = Font(bold=True, size=9)
        ws.merge_cells(f"A{total_row}:E{total_row}")

        total_values = {
            6: totals["tuition"],
            7: totals["adm_fees"],
            8: totals["total_bill"],
            9: totals["concession"],
            10: totals["net_bill"],
            12: totals["amt_due_todate"],
            13: totals["total_paid"],
            14: totals["balance"],
            16: (totals["pct_paid_net"] / 100) if totals["pct_paid_net"] else 0,
        }

        total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        total_font = Font(bold=True, size=9)

        for col_idx in range(1, 18):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = thin_border
            if col_idx in total_values:
                cell.value = total_values[col_idx]
                cell.alignment = Alignment(horizontal="right")
                if col_idx in currency_cols:
                    cell.number_format = money_fmt
                elif col_idx in pct_cols:
                    cell.number_format = "0.0%"

        # Column widths & freeze panes
        col_widths = [12, 28, 16, 14, 8, 12, 12, 12, 12, 12, 16, 14, 12, 12, 17, 16, 16]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = f"A{DATA_START}"

        # Stream response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_year = academic_year.name.replace(" ", "-").replace("/", "-").lower()
        filename = f"fees-sitrep-{safe_year}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_pdf(self, request, results, totals, academic_year):
        from ..utils.export_helpers import build_pdf_response

        headers = [
            "Student ID",
            "Student Name",
            "Grade",
            "Section",
            "Net Bill",
            "Total Paid",
            "Balance",
            "Status",
        ]
        rows = [
            [
                row["student_id"],
                row["student_name"],
                row["grade_level"],
                row["section"],
                row["net_bill"],
                row["total_paid"],
                row["balance"],
                row["status"],
            ]
            for row in results
        ]
        safe_year = academic_year.name.replace(" ", "-").replace("/", "-").lower()
        return build_pdf_response(
            request=request,
            filename=f"student-payment-summary-{safe_year}.pdf",
            title="Student Payment Summary",
            subtitle=f"Academic Year {academic_year.name} — {totals['student_count']} students",
            headers=headers,
            rows=rows,
        )
